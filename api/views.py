from datetime import timedelta, date
from decimal import Decimal
from math import ceil

import openpyxl
from dateutil.relativedelta import relativedelta
from django.db.models import ExpressionWrapper, F, DecimalField, Sum, Count, Value, Min, Max, Q
from django.db.models.functions import TruncDay, TruncWeek, TruncMonth, Coalesce
from django.utils import timezone
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.pagination import PageNumberPagination
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from user.models import User
from .filters import WarehouseProductFilter, OutcomeFilter, OrderFilter, IncomeFilter
from .models import ReportItem, Report
from .serializers import *


class CustomPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100


class StatusViewSet(viewsets.ModelViewSet):
    queryset = Status.objects.order_by('-id')
    serializer_class = StatusSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPagination
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['name', 'status', '_type', 'created', 'user']


class UnitTypeViewSet(viewsets.ModelViewSet):
    queryset = UnitType.objects.order_by('-id')
    serializer_class = UnitTypeSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPagination
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['name', 'status', 'created', 'user']


class ProductCategoryViewSet(viewsets.ModelViewSet):
    queryset = ProductCategory.objects.order_by('-id')
    serializer_class = ProductCategorySerializer
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPagination
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['name', 'status', 'created', 'user']


class ProductViewSet(viewsets.ModelViewSet):
    queryset = Product.objects.order_by('-id')
    serializer_class = ProductSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPagination
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['name', 'category', 'code', 'unit_type', 'status', 'created', 'user']


class WarehouseViewSet(viewsets.ModelViewSet):
    queryset = Warehouse.objects.order_by('-id')
    serializer_class = WarehouseSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPagination
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['name', 'created', 'user', 'responsible']


class WarehouseProductViewSet(viewsets.ModelViewSet):
    queryset = WarehouseProduct.objects.order_by('-id')
    serializer_class = WarehouseProductSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPagination
    filter_backends = [DjangoFilterBackend]
    filterset_class = WarehouseProductFilter


class IncomeViewSet(viewsets.ModelViewSet):
    queryset = Income.objects.order_by('-id')
    serializer_class = IncomeSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPagination
    filter_backends = [DjangoFilterBackend]
    filterset_class = IncomeFilter


class IncomeItemViewSet(viewsets.ModelViewSet):
    queryset = IncomeItem.objects.order_by('-id')
    serializer_class = IncomeItemSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPagination
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['income', 'product', 'status', 'user']


class OutcomeViewSet(viewsets.ModelViewSet):
    queryset = Outcome.objects.order_by('-id')
    serializer_class = OutcomeSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPagination
    filter_backends = [DjangoFilterBackend]
    filterset_class = OutcomeFilter

    @action(detail=False, methods=['get'], url_path='my')
    def my_outcomes(self, request):
        """Возвращает исходы (Outcome), где client = текущий пользователь"""
        user = request.user
        # Проверяем роль пользователя
        if not hasattr(user, 'role') or user.role != 'dealer':
            return Response(
                {"detail": "Доступ запрещён: только пользователи с ролью 'shop'."},
                status=status.HTTP_403_FORBIDDEN
            )

        queryset = self.get_queryset().filter(client=user)
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


class OutcomeItemViewSet(viewsets.ModelViewSet):
    queryset = OutcomeItem.objects.order_by('-id')
    serializer_class = OutcomeItemSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPagination
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['outcome', 'product', 'status', 'user']


class MovementViewSet(viewsets.ModelViewSet):
    queryset = Movement.objects.order_by('-id')
    serializer_class = MovementSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPagination
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['warehouse_from', 'warehouse_to', 'user', 'created', 'status']


class MovementItemViewSet(viewsets.ModelViewSet):
    queryset = MovementItem.objects.order_by('-id')
    serializer_class = MovementItemSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPagination
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['movement', 'product', 'user']


class OrderViewSet(viewsets.ModelViewSet):
    queryset = Order.objects.order_by('-id')
    serializer_class = OrderSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPagination
    filter_backends = [DjangoFilterBackend]
    filterset_class = OrderFilter

    @action(detail=False, methods=['get'], url_path='my')
    def my_orders(self, request):
        user = request.user
        # Проверяем роль пользователя
        if not hasattr(user, 'role') or user.role != 'provider':
            return Response(
                {"detail": "Доступ запрещён: только пользователи с ролью 'shop'."},
                status=status.HTTP_403_FORBIDDEN
            )

        queryset = self.get_queryset().filter(client=user)
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


class OrderItemViewSet(viewsets.ModelViewSet):
    queryset = OrderItem.objects.order_by('-id')
    serializer_class = OrderItemSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPagination
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['order', 'product', 'status']


class SalesVolumeView(APIView):
    """Общий объем продаж (по дням, неделям, месяцам)"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        group_by = request.query_params.get("group_by", "day")  # day|week|month
        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")
        warehouse_id = request.query_params.get("warehouse")
        client_id = request.query_params.get("client")
        product_id = request.query_params.get("product")
        status = request.query_params.get("status", Outcome.Status.finished)

        trunc = {"day": TruncDay, "week": TruncWeek, "month": TruncMonth}.get(group_by, TruncDay)

        qs = OutcomeItem.objects.select_related("outcome", "product")
        if status:
            qs = qs.filter(outcome__status=status)
        if date_from:
            qs = qs.filter(outcome__created__date__gte=date_from)
        if date_to:
            qs = qs.filter(outcome__created__date__lte=date_to)
        if warehouse_id:
            qs = qs.filter(outcome__warehouse_id=warehouse_id)
        if client_id:
            qs = qs.filter(outcome__client_id=client_id)
        if product_id:
            qs = qs.filter(product_id=product_id)

        amount_expr = ExpressionWrapper(
            F("count") * F("price"),
            output_field=DecimalField(max_digits=20, decimal_places=2)
        )

        zero_dec = Value(0, output_field=DecimalField(max_digits=20, decimal_places=2))

        agg = (
            qs.annotate(period=trunc("outcome__created"))
            .values("period")
            .annotate(
                total_qty=Coalesce(Sum("count"), 0),  # Int OK
                total_amount=Coalesce(Sum(amount_expr), zero_dec),  # <-- Decimal OK
                orders=Count("outcome_id", distinct=True),
            )
            .order_by("period")
        )

        data = [
            {
                "period": item["period"].date().isoformat() if hasattr(item["period"], "date") else item["period"],
                "total_qty": item["total_qty"],
                "total_amount": item["total_amount"],
                "orders": item["orders"],
            }
            for item in agg
        ]
        return Response({"group_by": group_by, "results": data})


class SalesVolumeCompareView(APIView):
    """Сравнение с предыдущими периодами"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        group_by = request.query_params.get("group_by", "day")  # day|week|month
        mode = request.query_params.get("mode", "prev")  # prev|yoy
        status = request.query_params.get("status", Outcome.Status.finished)

        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")

        # --- Определяем период по умолчанию (последние 30 дней)
        if not date_to or not date_from:
            now = timezone.now().date()
            date_to = date_to or now.isoformat()
            date_from = date_from or (now - timedelta(days=30)).isoformat()

        # Текущий диапазон
        cur_start = date_from
        cur_end = date_to

        # Сравнительный диапазон
        if mode == "yoy":
            # Тот же период год назад
            prev_start = (timezone.datetime.fromisoformat(cur_start) - relativedelta(years=1)).date().isoformat()
            prev_end = (timezone.datetime.fromisoformat(cur_end) - relativedelta(years=1)).date().isoformat()
        else:
            # Непосредственно предыдущий равный по длине период
            cur_len_days = (timezone.datetime.fromisoformat(cur_end).date()
                            - timezone.datetime.fromisoformat(cur_start).date()).days or 0
            prev_end_date = timezone.datetime.fromisoformat(cur_start).date() - timedelta(days=1)
            prev_start_date = prev_end_date - timedelta(days=cur_len_days)
            prev_start, prev_end = prev_start_date.isoformat(), prev_end_date.isoformat()

        trunc = {"day": TruncDay, "week": TruncWeek, "month": TruncMonth}.get(group_by, TruncDay)

        # Базовый QS + фильтры
        def base_qs():
            qs = OutcomeItem.objects.select_related("outcome", "product")
            if status:
                qs = qs.filter(outcome__status=status)
            warehouse = request.query_params.get("warehouse")
            client = request.query_params.get("client")
            product = request.query_params.get("product")
            if warehouse:
                qs = qs.filter(outcome__warehouse_id=warehouse)
            if client:
                qs = qs.filter(outcome__client_id=client)
            if product:
                qs = qs.filter(product_id=product)
            return qs

        amount_expr = ExpressionWrapper(
            F("count") * F("price"),
            output_field=DecimalField(max_digits=20, decimal_places=2)
        )
        zero_dec = Value(0, output_field=DecimalField(max_digits=20, decimal_places=2))

        def aggregate_range(start_date: str, end_date: str):
            qs = base_qs().filter(
                outcome__created__date__gte=start_date,
                outcome__created__date__lte=end_date
            )
            # Итоги
            totals = qs.aggregate(
                total_qty=Coalesce(Sum("count"), 0),
                total_amount=Coalesce(Sum(amount_expr), zero_dec),
                orders=Count("outcome_id", distinct=True),
            )
            # По корзинам
            buckets = (
                qs.annotate(period=trunc("outcome__created"))
                .values("period")
                .annotate(
                    total_qty=Coalesce(Sum("count"), 0),
                    total_amount=Coalesce(Sum(amount_expr), zero_dec),
                    orders=Count("outcome_id", distinct=True),
                )
                .order_by("period")
            )
            # нормализуем ключ периода в iso
            by_period = []
            for b in buckets:
                p = b["period"]
                key = p.date().isoformat() if hasattr(p, "date") else p
                by_period.append({
                    "period": key,
                    "total_qty": b["total_qty"],
                    "total_amount": b["total_amount"],
                    "orders": b["orders"],
                })
            return totals, by_period

        cur_totals, cur_periods = aggregate_range(cur_start, cur_end)
        prev_totals, prev_periods = aggregate_range(prev_start, prev_end)

        # Вычисляем дельты/проценты
        def diff(cur, prev):
            def pct(c, p):
                return float(c - p) / float(p) * 100.0 if p not in (0, None) else None

            return {
                "qty": {
                    "current": cur["total_qty"],
                    "previous": prev["total_qty"],
                    "delta": cur["total_qty"] - prev["total_qty"],
                    "pct": pct(cur["total_qty"], prev["total_qty"]),
                },
                "amount": {
                    "current": cur["total_amount"],
                    "previous": prev["total_amount"],
                    "delta": cur["total_amount"] - prev["total_amount"],
                    "pct": pct(cur["total_amount"], prev["total_amount"]),
                },
                "orders": {
                    "current": cur["orders"],
                    "previous": prev["orders"],
                    "delta": cur["orders"] - prev["orders"],
                    "pct": pct(cur["orders"], prev["orders"]),
                }
            }

        # Сшиваем корзины по ключу периода
        prev_map = {b["period"]: b for b in prev_periods}
        by_period = []
        for b in cur_periods:
            pkey = b["period"]
            prev_b = prev_map.get(pkey, {"total_qty": 0, "total_amount": 0, "orders": 0})

            def pct(c, p):
                return float(c - p) / float(p) * 100.0 if p not in (0, None) else None

            by_period.append({
                "period": pkey,
                "qty": {
                    "current": b["total_qty"],
                    "previous": prev_b["total_qty"],
                    "delta": b["total_qty"] - prev_b["total_qty"],
                    "pct": pct(b["total_qty"], prev_b["total_qty"]),
                },
                "amount": {
                    "current": b["total_amount"],
                    "previous": prev_b["total_amount"],
                    "delta": b["total_amount"] - prev_b["total_amount"],
                    "pct": pct(b["total_amount"], prev_b["total_amount"]),
                },
                "orders": {
                    "current": b["orders"],
                    "previous": prev_b["orders"],
                    "delta": b["orders"] - prev_b["orders"],
                    "pct": pct(b["orders"], prev_b["orders"]),
                },
            })

        result = {
            "group_by": group_by,
            "mode": mode,
            "current": {"date_from": cur_start, "date_to": cur_end},
            "previous": {"date_from": prev_start, "date_to": prev_end},
            "summary": diff(cur_totals, prev_totals),
            "by_period": by_period,
        }
        return Response(result)


class TopProductsView(APIView):
    """
    Самые продаваемые товары
    GET /api/reports/top-products/
        ?metric=amount|qty
        &limit=20
        &date_from=YYYY-MM-DD
        &date_to=YYYY-MM-DD
        &warehouse=ID
        &client=ID
        &status=finished|active|...
        &category=ID
        &product=ID   (можно передавать несколько: ?product=560&product=561)

    Возвращает топ товаров с полями:
    - product_id, product_name, unit_type, category_id, category_name
    - total_qty, total_amount, orders
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        metric = request.query_params.get("metric", "amount")  # amount|qty
        try:
            limit = int(request.query_params.get("limit", 20))
        except ValueError:
            limit = 20
        limit = max(1, min(limit, 500))  # ограничим разумно

        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")
        warehouse_id = request.query_params.get("warehouse")
        client_id = request.query_params.get("client")
        category_id = request.query_params.get("category")
        status = request.query_params.get("status", Outcome.Status.finished)

        qs = OutcomeItem.objects.select_related("outcome", "product", "product__category")

        # Фильтры
        if status:
            qs = qs.filter(outcome__status=status)
        if date_from:
            qs = qs.filter(outcome__created__date__gte=date_from)
        if date_to:
            qs = qs.filter(outcome__created__date__lte=date_to)
        if warehouse_id:
            qs = qs.filter(outcome__warehouse_id=warehouse_id)
        if client_id:
            qs = qs.filter(outcome__client_id=client_id)
        if category_id:
            qs = qs.filter(product__category_id=category_id)

        # фильтр по нескольким product (повторяющийся параметр в query string)
        product_ids = request.query_params.getlist("product")
        if product_ids:
            qs = qs.filter(product_id__in=product_ids)

        # Агрегации
        amount_expr = ExpressionWrapper(
            F("count") * F("price"),
            output_field=DecimalField(max_digits=20, decimal_places=2),
        )
        zero_dec = Value(0, output_field=DecimalField(max_digits=20, decimal_places=2))

        grouped = (
            qs.values(
                "product_id",
                "product__name",
                "product__unit_type",
                "product__category_id",
                "product__category__name",
            )
            .annotate(
                total_qty=Coalesce(Sum("count"), 0),
                total_amount=Coalesce(Sum(amount_expr), zero_dec),
                orders=Count("outcome_id", distinct=True),
            )
        )

        if metric == "qty":
            grouped = grouped.order_by("-total_qty", "product__name")
        else:  # amount
            grouped = grouped.order_by("-total_amount", "product__name")

        data = [
            {
                "product_id": r["product_id"],
                "product_name": r["product__name"],
                "unit_type": r["product__unit_type"],
                "category_id": r["product__category_id"],
                "category_name": r["product__category__name"],
                "total_qty": r["total_qty"],
                "total_amount": r["total_amount"],
                "orders": r["orders"],
            }
            for r in grouped[:limit]
        ]

        # опционально — добавить номер в рейтинге
        for i, item in enumerate(data, start=1):
            item["rank"] = i

        return Response({
            "metric": metric,
            "limit": limit,
            "count": len(data),
            "results": data
        })


class LeastPopularProductsView(APIView):
    """
    Наименее популярные товары
    GET /api/reports/least-popular-products/
        ?metric=amount|qty
        &limit=20
        &date_from=YYYY-MM-DD
        &date_to=YYYY-MM-DD
        &warehouse=ID
        &client=ID
        &status=finished|active|...
        &category=ID
        &product=ID

    Возвращает наименее популярные товары:
    - product_id, product_name, unit_type, category_id, category_name
    - total_qty, total_amount, orders
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        metric = request.query_params.get("metric", "amount")
        try:
            limit = int(request.query_params.get("limit", 20))
        except ValueError:
            limit = 20
        limit = max(1, min(limit, 500))

        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")
        warehouse_id = request.query_params.get("warehouse")
        client_id = request.query_params.get("client")
        category_id = request.query_params.get("category")
        status = request.query_params.get("status", Outcome.Status.finished)

        qs = OutcomeItem.objects.select_related("outcome", "product", "product__category")

        # Фильтры
        if status:
            qs = qs.filter(outcome__status=status)
        if date_from:
            qs = qs.filter(outcome__created__date__gte=date_from)
        if date_to:
            qs = qs.filter(outcome__created__date__lte=date_to)
        if warehouse_id:
            qs = qs.filter(outcome__warehouse_id=warehouse_id)
        if client_id:
            qs = qs.filter(outcome__client_id=client_id)
        if category_id:
            qs = qs.filter(product__category_id=category_id)
        product_ids = request.query_params.getlist("product")
        if product_ids:
            qs = qs.filter(product_id__in=product_ids)

        # Агрегации
        amount_expr = ExpressionWrapper(
            F("count") * F("price"),
            output_field=DecimalField(max_digits=20, decimal_places=2),
        )
        zero_dec = Value(0, output_field=DecimalField(max_digits=20, decimal_places=2))

        grouped = (
            qs.values(
                "product_id",
                "product__name",
                "product__unit_type",
                "product__category_id",
                "product__category__name",
            )
            .annotate(
                total_qty=Coalesce(Sum("count"), 0),
                total_amount=Coalesce(Sum(amount_expr), zero_dec),
                orders=Count("outcome_id", distinct=True),
            )
        )

        # Сортировка: наоборот (от наименьшего к большему)
        if metric == "qty":
            grouped = grouped.order_by("total_qty", "product__name")
        else:
            grouped = grouped.order_by("total_amount", "product__name")

        data = []
        for i, r in enumerate(grouped[:limit], start=1):
            data.append({
                "rank": i,
                "product_id": r["product_id"],
                "product_name": r["product__name"],
                "unit_type": r["product__unit_type"],
                "category_id": r["product__category_id"],
                "category_name": r["product__category__name"],
                "total_qty": r["total_qty"],
                "total_amount": r["total_amount"],
                "orders": r["orders"],
            })

        return Response({
            "metric": metric,
            "limit": limit,
            "count": len(data),
            "results": data
        })


class DealersSalesView(APIView):
    """
    Список всех дилеров с объемом продаж
    GET /api/reports/dealers-sales/
        ?date_from=YYYY-MM-DD
        &date_to=YYYY-MM-DD
        &warehouse=ID
        &status=finished|active|...
        &client=1&client=2     (несколько)
        &order_by=amount|qty|orders|name|last_sale
        &direction=desc|asc
        &limit=100

    Возвращает список дилеров (users) с объёмом продаж.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")
        warehouse_id = request.query_params.get("warehouse")
        status = request.query_params.get("status", Outcome.Status.finished)
        client_ids = request.query_params.getlist("client")

        order_by = request.query_params.get("order_by", "amount")
        direction = request.query_params.get("direction", "desc")
        try:
            limit = int(request.query_params.get("limit", 100))
        except ValueError:
            limit = 100
        limit = max(1, min(limit, 1000))

        qs = OutcomeItem.objects.select_related(
            "outcome", "outcome__client"
        )

        # Фильтры
        if status:
            qs = qs.filter(outcome__status=status)
        if date_from:
            qs = qs.filter(outcome__created__date__gte=date_from)
        if date_to:
            qs = qs.filter(outcome__created__date__lte=date_to)
        if warehouse_id:
            qs = qs.filter(outcome__warehouse_id=warehouse_id)
        if client_ids:
            qs = qs.filter(outcome__client_id__in=client_ids)

        # Агрегации
        amount_expr = ExpressionWrapper(
            F("count") * F("price"),
            output_field=DecimalField(max_digits=20, decimal_places=2),
        )
        zero_dec = Value(0, output_field=DecimalField(max_digits=20, decimal_places=2))

        grouped = (
            qs.values(
                "outcome__client_id",
                "outcome__client__first_name",
                "outcome__client__last_name",
            )
            .annotate(
                total_qty=Coalesce(Sum("count"), 0),
                total_amount=Coalesce(Sum(amount_expr), zero_dec),
                orders=Count("outcome_id", distinct=True),
                first_sale=Min("outcome__created"),
                last_sale=Max("outcome__created"),
                warehouses=Count("outcome__warehouse_id", distinct=True),
            )
        )

        # Сортировка
        order_map = {
            "amount": "total_amount",
            "qty": "total_qty",
            "orders": "orders",
            "name": "outcome__client__first_name",
            "last_sale": "last_sale",
        }
        key = order_map.get(order_by, "total_amount")
        if direction == "asc":
            grouped = grouped.order_by(key)
        else:
            grouped = grouped.order_by(f"-{key}")

        # Формирование ответа + avg_check
        results = []
        for row in grouped[:limit]:
            client_id = row["outcome__client_id"]
            fname = row["outcome__client__first_name"] or ""
            lname = row["outcome__client__last_name"] or ""
            full_name = (fname + " " + lname).strip()
            display_name = full_name or f"User {client_id}"

            orders = row["orders"] or 0
            total_amount = row["total_amount"]
            avg_check = (total_amount / orders) if orders else None

            results.append({
                "client_id": client_id,
                "client_full_name": full_name,
                "display_name": display_name,
                "total_qty": row["total_qty"],
                "total_amount": total_amount,
                "orders": orders,
                "avg_check": avg_check,
                "first_sale": row["first_sale"],
                "last_sale": row["last_sale"],
                "warehouses": row["warehouses"],
            })

        return Response({
            "count": len(results),
            "order_by": order_by,
            "direction": direction,
            "results": results,
        })


class DealersCompareView(APIView):
    """
    Сравнение активности разных дилеров
    GET /api/reports/dealers-compare/
        ?metric=amount|qty|orders
        &date_from=YYYY-MM-DD
        &date_to=YYYY-MM-DD
        &warehouse=ID
        &status=finished|active|...
        &client=1&client=7...
        &limit=50
        &group_by=day|week|month

    Возвращает:
    - results: по каждому дилеру total_qty, total_amount, orders, доли (%) и индекс 0..100 по выбранной метрике
    - timeseries: (если задан group_by) временные ряды по каждому дилеру
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        metric = request.query_params.get("metric", "amount")  # amount|qty|orders
        group_by = request.query_params.get("group_by")  # None|day|week|month

        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")
        warehouse_id = request.query_params.get("warehouse")
        status = request.query_params.get("status", Outcome.Status.finished)
        client_ids = request.query_params.getlist("client")

        try:
            limit = int(request.query_params.get("limit", 50))
        except ValueError:
            limit = 50
        limit = max(1, min(limit, 500))

        # База + фильтры
        qs = OutcomeItem.objects.select_related("outcome", "outcome__client")
        if status:
            qs = qs.filter(outcome__status=status)
        if date_from:
            qs = qs.filter(outcome__created__date__gte=date_from)
        if date_to:
            qs = qs.filter(outcome__created__date__lte=date_to)
        if warehouse_id:
            qs = qs.filter(outcome__warehouse_id=warehouse_id)
        if client_ids:
            qs = qs.filter(outcome__client_id__in=client_ids)

        # Выражения ORM
        amount_expr = ExpressionWrapper(
            F("count") * F("price"),
            output_field=DecimalField(max_digits=20, decimal_places=2),
        )
        zero_dec = Value(0, output_field=DecimalField(max_digits=20, decimal_places=2))  # ТОЛЬКО для ORM/Coalesce

        # Агрегация по дилерам
        grouped = (
            qs.values(
                "outcome__client_id",
                "outcome__client__first_name",
                "outcome__client__last_name",
            )
            .annotate(
                total_qty=Coalesce(Sum("count"), 0),
                total_amount=Coalesce(Sum(amount_expr), zero_dec),
                orders=Count("outcome_id", distinct=True),
            )
        )

        order_key = {
            "qty": "-total_qty",
            "orders": "-orders",
            "amount": "-total_amount",
        }.get(metric, "-total_amount")

        # Применяем limit только если явно не задан список client
        if client_ids:
            grouped = grouped.order_by(order_key)
            base_rows = list(grouped)
        else:
            base_rows = list(grouped.order_by(order_key)[:limit])

        # Тоталы (для расчёта долей) считаем по выбранному набору дилеров
        chosen_client_ids = [r["outcome__client_id"] for r in base_rows]
        totals_qs = qs.filter(outcome__client_id__in=chosen_client_ids).aggregate(
            sum_qty=Coalesce(Sum("count"), 0),
            sum_amount=Coalesce(Sum(amount_expr), zero_dec),
            sum_orders=Count("outcome_id", distinct=True),
        )
        sum_qty = totals_qs["sum_qty"] or 0
        sum_amount = totals_qs["sum_amount"] or Decimal("0")
        sum_orders = totals_qs["sum_orders"] or 0

        # Максимумы для индекса 0..100 (используем чистые Decimal/ints)
        max_qty = 0
        max_amount = Decimal("0")
        max_orders = 0
        for r in base_rows:
            if r["total_qty"] > max_qty:
                max_qty = r["total_qty"]
            if r["total_amount"] > max_amount:
                max_amount = r["total_amount"]
            if r["orders"] > max_orders:
                max_orders = r["orders"]

        # Основная таблица результатов
        results = []
        for r in base_rows:
            cid = r["outcome__client_id"]
            fname = r["outcome__client__first_name"] or ""
            lname = r["outcome__client__last_name"] or ""
            full_name = (fname + " " + lname).strip()
            display_name = full_name or f"User {cid}"

            tq = r["total_qty"]
            ta = r["total_amount"]
            ords = r["orders"]

            # доли
            share_qty = float(tq) / float(sum_qty) * 100.0 if sum_qty else None
            share_amount = float(ta) / float(sum_amount) * 100.0 if sum_amount and float(sum_amount) != 0.0 else None
            share_orders = float(ords) / float(sum_orders) * 100.0 if sum_orders else None

            # индекс 0..100 по выбранной метрике
            if metric == "qty":
                idx = (float(tq) / float(max_qty) * 100.0) if max_qty else None
            elif metric == "orders":
                idx = (float(ords) / float(max_orders) * 100.0) if max_orders else None
            else:  # amount
                idx = (float(ta) / float(max_amount) * 100.0) if max_amount and float(max_amount) != 0.0 else None

            results.append({
                "client_id": cid,
                "display_name": display_name,
                "total_qty": tq,
                "total_amount": ta,
                "orders": ords,
                "share_qty_pct": share_qty,
                "share_amount_pct": share_amount,
                "share_orders_pct": share_orders,
                "index_0_100": idx,
            })

        # Таймсерия по каждому дилеру (по запросу)
        timeseries = None
        if group_by in {"day", "week", "month"} and chosen_client_ids:
            trunc = {"day": TruncDay, "week": TruncWeek, "month": TruncMonth}[group_by]
            buckets = (
                qs.filter(outcome__client_id__in=chosen_client_ids)
                .annotate(period=trunc("outcome__created"))
                .values("period", "outcome__client_id")
                .annotate(
                    total_qty=Coalesce(Sum("count"), 0),
                    total_amount=Coalesce(Sum(amount_expr), zero_dec),
                    orders=Count("outcome_id", distinct=True),
                )
                .order_by("period")
            )
            ts_map = {cid: [] for cid in chosen_client_ids}
            for b in buckets:
                p = b["period"]
                period_key = p.date().isoformat() if hasattr(p, "date") else p
                ts_map[b["outcome__client_id"]].append({
                    "period": period_key,
                    "total_qty": b["total_qty"],
                    "total_amount": b["total_amount"],
                    "orders": b["orders"],
                })
            timeseries = [{"client_id": cid, "series": ts_map.get(cid, [])} for cid in chosen_client_ids]

        return Response({
            "metric": metric,
            "date_from": date_from,
            "date_to": date_to,
            "limit": None if client_ids else limit,
            "results": results,
            "timeseries": timeseries,
        })


class DealerAvgCheckView(APIView):
    """
    Средний чек по дилеру
    GET /api/reports/dealer-avg-check/
        ?date_from=YYYY-MM-DD
        &date_to=YYYY-MM-DD
        &warehouse=ID
        &status=finished|active|...
        &client=1&client=7...       # можно несколько
        &order_by=avg_check|amount|orders|qty|name|last_sale
        &direction=desc|asc
        &limit=100

    Возвращает по каждому дилеру:
      - total_amount, orders, avg_check (= total_amount / orders), total_qty
      - first_sale, last_sale, warehouses (в скольких складах были продажи)
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")
        warehouse_id = request.query_params.get("warehouse")
        status = request.query_params.get("status", Outcome.Status.finished)
        client_ids = request.query_params.getlist("client")

        order_by = request.query_params.get("order_by", "avg_check")
        direction = request.query_params.get("direction", "desc")
        try:
            limit = int(request.query_params.get("limit", 100))
        except ValueError:
            limit = 100
        limit = max(1, min(limit, 1000))

        # База + фильтры
        qs = OutcomeItem.objects.select_related("outcome", "outcome__client")
        if status:
            qs = qs.filter(outcome__status=status)
        if date_from:
            qs = qs.filter(outcome__created__date__gte=date_from)
        if date_to:
            qs = qs.filter(outcome__created__date__lte=date_to)
        if warehouse_id:
            qs = qs.filter(outcome__warehouse_id=warehouse_id)
        if client_ids:
            qs = qs.filter(outcome__client_id__in=client_ids)

        # ORM выражения
        amount_expr = ExpressionWrapper(
            F("count") * F("price"),
            output_field=DecimalField(max_digits=20, decimal_places=2),
        )
        zero_dec = Value(0, output_field=DecimalField(max_digits=20, decimal_places=2))

        grouped = (
            qs.values(
                "outcome__client_id",
                "outcome__client__first_name",
                "outcome__client__last_name",
            )
            .annotate(
                total_qty=Coalesce(Sum("count"), 0),
                total_amount=Coalesce(Sum(amount_expr), zero_dec),
                orders=Count("outcome_id", distinct=True),
                first_sale=Min("outcome__created"),
                last_sale=Max("outcome__created"),
                warehouses=Count("outcome__warehouse_id", distinct=True),
            )
        )

        # Формируем список и считаем avg_check в Python (избегаем деления в БД)
        rows = []
        for r in grouped:
            orders = r["orders"] or 0
            total_amount = r["total_amount"] or Decimal("0")
            avg_check = (total_amount / orders) if orders else None

            cid = r["outcome__client_id"]
            fname = r["outcome__client__first_name"] or ""
            lname = r["outcome__client__last_name"] or ""
            full_name = (fname + " " + lname).strip()
            display_name = full_name or f"User {cid}"

            rows.append({
                "client_id": cid,
                "display_name": display_name,
                "total_amount": total_amount,
                "orders": orders,
                "avg_check": avg_check,
                "total_qty": r["total_qty"],
                "first_sale": r["first_sale"],
                "last_sale": r["last_sale"],
                "warehouses": r["warehouses"],
            })

        # Сортировка
        key_map = {
            "avg_check": (lambda x: (x["avg_check"] is None, x["avg_check"] or Decimal("0"))),
            "amount": (lambda x: x["total_amount"]),
            "orders": (lambda x: x["orders"]),
            "qty": (lambda x: x["total_qty"]),
            "name": (lambda x: x["display_name"].lower()),
            "last_sale": (lambda x: x["last_sale"] or ""),
        }
        key_fn = key_map.get(order_by, key_map["avg_check"])
        rows.sort(key=key_fn, reverse=(direction != "asc"))

        # Ограничение
        if not client_ids:
            rows = rows[:limit]

        return Response({
            "order_by": order_by,
            "direction": direction,
            "limit": None if client_ids else limit,
            "count": len(rows),
            "results": rows,
        })


class OrdersAndReturnsView(APIView):
    """
    Количество заказов и возвратов
    GET /api/reports/orders-and-returns/
        ?group_by=day|week|month
        &date_from=YYYY-MM-DD
        &date_to=YYYY-MM-DD
        &client=ID                 # применится и к заказам, и к возвратам
        &warehouse=ID              # применится только к возвратам (Outcome.warehouse)
        &include_returns=orders|outcomes|both   # по умолчанию both

    Возвращает по каждому периоду:
      {
        "period": "2025-08-01",
        "orders_total": 35,
        "orders_cancelled": 2,
        "returns_outcomes_cancelled": 1,
        "returns_total": 3,
        "orders_cancel_rate_pct": 5.71,
        "returns_share_pct": 7.89
      }
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        group_by = request.query_params.get("group_by", "day")  # day|week|month
        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")
        client_id = request.query_params.get("client")
        warehouse_id = request.query_params.get("warehouse")
        include_returns = request.query_params.get("include_returns", "both")  # orders|outcomes|both

        trunc = {"day": TruncDay, "week": TruncWeek, "month": TruncMonth}.get(group_by, TruncDay)

        # ---- Заказы (Order)
        orders_qs = Order.objects.all()
        if date_from:
            orders_qs = orders_qs.filter(created__date__gte=date_from)
        if date_to:
            orders_qs = orders_qs.filter(created__date__lte=date_to)
        if client_id:
            orders_qs = orders_qs.filter(client_id=client_id)

        orders_bucketed = (
            orders_qs.annotate(period=trunc("created"))
            .values("period")
            .annotate(
                orders_total=Coalesce(Count("id"), 0),
                orders_cancelled=Coalesce(Count("id", filter=Q(status=Order.Status.cancelled)), 0),
            )
            .order_by("period")
        )

        # ---- Возвраты/отмены отгрузок (Outcome с cancelled)
        returns_bucketed = []
        if include_returns in ("outcomes", "both"):
            out_qs = Outcome.objects.filter(status=Outcome.Status.cancelled)
            if date_from:
                out_qs = out_qs.filter(created__date__gte=date_from)
            if date_to:
                out_qs = out_qs.filter(created__date__lte=date_to)
            if client_id:
                out_qs = out_qs.filter(client_id=client_id)
            if warehouse_id:
                out_qs = out_qs.filter(warehouse_id=warehouse_id)

            returns_bucketed = (
                out_qs.annotate(period=trunc("created"))
                .values("period")
                .annotate(returns_outcomes_cancelled=Coalesce(Count("id"), 0))
                .order_by("period")
            )

        # ---- Возвраты по отменённым заказам (если нужно учитывать как «returns»)
        orders_returns_bucketed = []
        if include_returns in ("orders", "both"):
            or_qs = orders_qs.filter(status=Order.Status.cancelled)
            orders_returns_bucketed = (
                or_qs.annotate(period=trunc("created"))
                .values("period")
                .annotate(orders_cancelled_as_returns=Coalesce(Count("id"), 0))
                .order_by("period")
            )

        # ---- Сшивка по периодам
        # Собираем все ключи периодов из трёх источников
        bucket_map = {}

        for row in orders_bucketed:
            p = row["period"]
            key = p.date().isoformat() if hasattr(p, "date") else p
            bucket_map[key] = {
                "period": key,
                "orders_total": row["orders_total"],
                "orders_cancelled": row["orders_cancelled"],
                "returns_outcomes_cancelled": 0,
                "returns_total": 0,
                "orders_cancel_rate_pct": None,
                "returns_share_pct": None,
            }

        for row in returns_bucketed:
            p = row["period"]
            key = p.date().isoformat() if hasattr(p, "date") else p
            bucket_map.setdefault(key, {
                "period": key,
                "orders_total": 0,
                "orders_cancelled": 0,
                "returns_outcomes_cancelled": 0,
                "returns_total": 0,
                "orders_cancel_rate_pct": None,
                "returns_share_pct": None,
            })
            bucket_map[key]["returns_outcomes_cancelled"] = row["returns_outcomes_cancelled"]

        for row in orders_returns_bucketed:
            p = row["period"]
            key = p.date().isoformat() if hasattr(p, "date") else p
            bucket_map.setdefault(key, {
                "period": key,
                "orders_total": 0,
                "orders_cancelled": 0,
                "returns_outcomes_cancelled": 0,
                "returns_total": 0,
                "orders_cancel_rate_pct": None,
                "returns_share_pct": None,
            })
            # Учтём эти отмены и как отменённые заказы, и как компонент возвратов
            bucket_map[key]["orders_cancelled"] += row["orders_cancelled_as_returns"]

        # ---- Финальные метрики по каждому периоду
        results = []
        for key in sorted(bucket_map.keys()):
            b = bucket_map[key]
            # returns_total = отменённые заказы + отменённые расходы (в зависимости от include_returns)
            returns_total = 0
            if include_returns in ("orders", "both"):
                returns_total += b["orders_cancelled"]
            if include_returns in ("outcomes", "both"):
                returns_total += b["returns_outcomes_cancelled"]
            b["returns_total"] = returns_total

            orders_total = b["orders_total"]
            if orders_total:
                b["orders_cancel_rate_pct"] = round(b["orders_cancelled"] * 100.0 / orders_total, 2)
                b["returns_share_pct"] = round(returns_total * 100.0 / orders_total, 2)
            else:
                b["orders_cancel_rate_pct"] = None
                b["returns_share_pct"] = None

            results.append(b)

        return Response({
            "group_by": group_by,
            "filters": {
                "date_from": date_from,
                "date_to": date_to,
                "client": client_id,
                "warehouse": warehouse_id,
                "include_returns": include_returns,
            },
            "results": results,
        })


class SalesGeographyView(APIView):
    """
    География продаж
    GET /api/reports/sales-geography/
        ?date_from=YYYY-MM-DD
        &date_to=YYYY-MM-DD
        &client=ID
        &status=finished|active|...

    Возвращает продажи в разрезе складов (география).
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")
        client_id = request.query_params.get("client")
        status = request.query_params.get("status", Outcome.Status.finished)

        qs = OutcomeItem.objects.select_related("outcome", "outcome__warehouse")

        # Фильтры
        if status:
            qs = qs.filter(outcome__status=status)
        if date_from:
            qs = qs.filter(outcome__created__date__gte=date_from)
        if date_to:
            qs = qs.filter(outcome__created__date__lte=date_to)
        if client_id:
            qs = qs.filter(outcome__client_id=client_id)

        # Выражения ORM
        amount_expr = ExpressionWrapper(
            F("count") * F("price"),
            output_field=DecimalField(max_digits=20, decimal_places=2),
        )
        zero_dec = Value(0, output_field=DecimalField(max_digits=20, decimal_places=2))

        grouped = (
            qs.values("outcome__warehouse_id", "outcome__warehouse__name")
            .annotate(
                total_qty=Coalesce(Sum("count"), 0),
                total_amount=Coalesce(Sum(amount_expr), zero_dec),
                orders=Count("outcome_id", distinct=True),
            )
            .order_by("-total_amount")
        )

        # Общие суммы для долей
        totals = qs.aggregate(
            sum_qty=Coalesce(Sum("count"), 0),
            sum_amount=Coalesce(Sum(amount_expr), zero_dec),
            sum_orders=Count("outcome_id", distinct=True),
        )
        sum_qty = totals["sum_qty"] or 0
        sum_amount = totals["sum_amount"] or Decimal("0")
        sum_orders = totals["sum_orders"] or 0

        results = []
        for r in grouped:
            share_amount = float(r["total_amount"]) / float(sum_amount) * 100.0 if sum_amount else None
            share_qty = float(r["total_qty"]) / float(sum_qty) * 100.0 if sum_qty else None
            share_orders = float(r["orders"]) / float(sum_orders) * 100.0 if sum_orders else None

            results.append({
                "warehouse_id": r["outcome__warehouse_id"],
                "warehouse_name": r["outcome__warehouse__name"],
                "total_qty": r["total_qty"],
                "total_amount": r["total_amount"],
                "orders": r["orders"],
                "share_amount_pct": share_amount,
                "share_qty_pct": share_qty,
                "share_orders_pct": share_orders,
            })

        return Response({
            "filters": {
                "date_from": date_from,
                "date_to": date_to,
                "client": client_id,
                "status": status,
            },
            "totals": {
                "sum_qty": sum_qty,
                "sum_amount": sum_amount,
                "sum_orders": sum_orders,
            },
            "results": results,
        })


class TopCategoriesView(APIView):
    """
    Какие категории товаров продаются лучше всего
    GET /api/reports/top-categories/
        ?date_from=YYYY-MM-DD
        &date_to=YYYY-MM-DD
        &warehouse=ID
        &client=ID
        &status=finished|active|...
        &metric=amount|qty|orders
        &limit=20

    Возвращает продажи по категориям товаров.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")
        warehouse_id = request.query_params.get("warehouse")
        client_id = request.query_params.get("client")
        status = request.query_params.get("status", Outcome.Status.finished)
        metric = request.query_params.get("metric", "amount")  # amount|qty|orders
        try:
            limit = int(request.query_params.get("limit", 20))
        except ValueError:
            limit = 20
        limit = max(1, min(limit, 500))

        qs = OutcomeItem.objects.select_related("outcome", "product", "product__category")

        # Фильтры
        if status:
            qs = qs.filter(outcome__status=status)
        if date_from:
            qs = qs.filter(outcome__created__date__gte=date_from)
        if date_to:
            qs = qs.filter(outcome__created__date__lte=date_to)
        if warehouse_id:
            qs = qs.filter(outcome__warehouse_id=warehouse_id)
        if client_id:
            qs = qs.filter(outcome__client_id=client_id)

        # ORM выражения
        amount_expr = ExpressionWrapper(
            F("count") * F("price"),
            output_field=DecimalField(max_digits=20, decimal_places=2),
        )
        zero_dec = Value(0, output_field=DecimalField(max_digits=20, decimal_places=2))

        grouped = (
            qs.values("product__category_id", "product__category__name")
            .annotate(
                total_qty=Coalesce(Sum("count"), 0),
                total_amount=Coalesce(Sum(amount_expr), zero_dec),
                orders=Count("outcome_id", distinct=True),
            )
        )

        # Сортировка
        order_key = {
            "qty": "-total_qty",
            "orders": "-orders",
            "amount": "-total_amount",
        }.get(metric, "-total_amount")
        grouped = grouped.order_by(order_key)

        # Общие суммы для долей
        totals = qs.aggregate(
            sum_qty=Coalesce(Sum("count"), 0),
            sum_amount=Coalesce(Sum(amount_expr), zero_dec),
            sum_orders=Count("outcome_id", distinct=True),
        )
        sum_qty = totals["sum_qty"] or 0
        sum_amount = totals["sum_amount"] or Decimal("0")
        sum_orders = totals["sum_orders"] or 0

        # Формируем список
        results = []
        for i, r in enumerate(grouped[:limit], start=1):
            share_amount = float(r["total_amount"]) / float(sum_amount) * 100.0 if sum_amount else None
            share_qty = float(r["total_qty"]) / float(sum_qty) * 100.0 if sum_qty else None
            share_orders = float(r["orders"]) / float(sum_orders) * 100.0 if sum_orders else None

            results.append({
                "rank": i,
                "category_id": r["product__category_id"],
                "category_name": r["product__category__name"],
                "total_qty": r["total_qty"],
                "total_amount": r["total_amount"],
                "orders": r["orders"],
                "share_amount_pct": share_amount,
                "share_qty_pct": share_qty,
                "share_orders_pct": share_orders,
            })

        return Response({
            "metric": metric,
            "filters": {
                "date_from": date_from,
                "date_to": date_to,
                "warehouse": warehouse_id,
                "client": client_id,
                "status": status,
            },
            "totals": {
                "sum_qty": sum_qty,
                "sum_amount": sum_amount,
                "sum_orders": sum_orders,
            },
            "results": results,
        })


class AssortmentStructureView(APIView):
    """
    Структура ассортимента и его эффективность
    GET /api/reports/assortment-structure/
        ?level=category|product          # уровень агрегации (по умолчанию: category)
        &metric=amount|qty|orders        # чем сортировать (по умолчанию amount)
        &limit=100                       # TOP-N строк (если нужно)
        &date_from=YYYY-MM-DD
        &date_to=YYYY-MM-DD
        &warehouse=ID
        &client=ID
        &status=finished|active|...

    Возвращает структуру ассортимента с метриками эффективности и ABC-классификацией.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        level = request.query_params.get("level", "category")  # category|product
        metric = request.query_params.get("metric", "amount")  # amount|qty|orders
        try:
            limit = int(request.query_params.get("limit", 100))
        except ValueError:
            limit = 100
        limit = max(1, min(limit, 1000))

        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")
        warehouse_id = request.query_params.get("warehouse")
        client_id = request.query_params.get("client")
        status = request.query_params.get("status", Outcome.Status.finished)

        qs = OutcomeItem.objects.select_related("outcome", "product", "product__category")

        # Фильтры
        if status:
            qs = qs.filter(outcome__status=status)
        if date_from:
            qs = qs.filter(outcome__created__date__gte=date_from)
        if date_to:
            qs = qs.filter(outcome__created__date__lte=date_to)
        if warehouse_id:
            qs = qs.filter(outcome__warehouse_id=warehouse_id)
        if client_id:
            qs = qs.filter(outcome__client_id=client_id)

        # ORM-выражения
        amount_expr = ExpressionWrapper(
            F("count") * F("price"),
            output_field=DecimalField(max_digits=20, decimal_places=2),
        )
        zero_dec = Value(0, output_field=DecimalField(max_digits=20, decimal_places=2))

        # Выбор полей группировки
        if level == "product":
            group_fields = (
                "product_id",
                "product__name",
                "product__unit_type",
                "product__category_id",
                "product__category__name",
            )
            result_keys = ("product_id", "product__name", "product__unit_type",
                           "product__category_id", "product__category__name")
        else:  # category (по умолчанию)
            group_fields = ("product__category_id", "product__category__name")
            result_keys = ("product__category_id", "product__category__name")

        grouped = (
            qs.values(*group_fields)
            .annotate(
                total_qty=Coalesce(Sum("count"), 0),
                total_amount=Coalesce(Sum(amount_expr), zero_dec),
                orders=Count("outcome_id", distinct=True),
            )
        )

        # Сортировка
        order_key = {
            "qty": "-total_qty",
            "orders": "-orders",
            "amount": "-total_amount",
        }.get(metric, "-total_amount")
        grouped = grouped.order_by(order_key)

        # Общие суммы для долей
        totals = qs.aggregate(
            sum_qty=Coalesce(Sum("count"), 0),
            sum_amount=Coalesce(Sum(amount_expr), zero_dec),
            sum_orders=Count("outcome_id", distinct=True),
        )
        sum_qty = totals["sum_qty"] or 0
        sum_amount = totals["sum_amount"] or Decimal("0")
        sum_orders = totals["sum_orders"] or 0

        # Собираем строки и считаем производные метрики в Python
        raw_rows = list(grouped[:limit]) if limit else list(grouped)

        # Для ABC нужна сортировка по total_amount (DESC) и кумулятивная доля
        rows_for_abc = sorted(raw_rows, key=lambda r: r["total_amount"], reverse=True)
        cum = Decimal("0")
        abc_map = {}
        if sum_amount and float(sum_amount) != 0.0:
            for r in rows_for_abc:
                key = self._key_from_row(level, r)
                cum += r["total_amount"]
                share = (cum / sum_amount) * Decimal("100")
                if share <= 80:
                    abc_map[key] = "A"
                elif share <= 95:
                    abc_map[key] = "B"
                else:
                    abc_map[key] = "C"

        # Формируем финальные строки
        results = []
        for i, r in enumerate(raw_rows, start=1):
            total_qty = r["total_qty"]
            total_amount = r["total_amount"]
            orders_cnt = r["orders"]

            avg_price = (total_amount / total_qty) if total_qty else None
            avg_qty_per_order = (total_qty / orders_cnt) if orders_cnt else None

            share_amount = float(total_amount) / float(sum_amount) * 100.0 if sum_amount else None
            share_qty = float(total_qty) / float(sum_qty) * 100.0 if sum_qty else None
            share_orders = float(orders_cnt) / float(sum_orders) * 100.0 if sum_orders else None

            base = {
                "rank": i,
                "total_qty": total_qty,
                "total_amount": total_amount,
                "orders": orders_cnt,
                "avg_price": avg_price,
                "avg_qty_per_order": avg_qty_per_order,
                "share_amount_pct": share_amount,
                "share_qty_pct": share_qty,
                "share_orders_pct": share_orders,
            }

            # ключ и наименования зависят от уровня
            if level == "product":
                base.update({
                    "product_id": r["product_id"],
                    "product_name": r["product__name"],
                    "unit_type": r["product__unit_type"],
                    "category_id": r["product__category_id"],
                    "category_name": r["product__category__name"],
                })
                abc_key = ("product", r["product_id"])
            else:
                base.update({
                    "category_id": r["product__category_id"],
                    "category_name": r["product__category__name"],
                })
                abc_key = ("category", r["product__category_id"])

            base["abc_class"] = abc_map.get(abc_key)  # может быть None, если sum_amount=0
            results.append(base)

        return Response({
            "level": level,
            "metric": metric,
            "filters": {
                "date_from": date_from,
                "date_to": date_to,
                "warehouse": warehouse_id,
                "client": client_id,
                "status": status,
            },
            "totals": {
                "sum_qty": sum_qty,
                "sum_amount": sum_amount,
                "sum_orders": sum_orders,
            },
            "results": results,
        })

    @staticmethod
    def _key_from_row(level, r):
        if level == "product":
            return ("product", r["product_id"])
        return ("category", r["product__category_id"])


class CentralStockView(APIView):
    """
    Общие остатки на центральном складе
    GET /api/reports/central-stock/
        ?warehouse=ID                # если не передан — пытаемся найти "Центральный" по имени
        &category=ID                 # фильтр по категории товара
        &product=ID                  # один или несколько ?product=560&product=561
        &include_zero=false|true     # включать позиции с нулевым/отрицательным остатком (по умолчанию false)
        &group_by=product|category   # разрез отчёта (по умолчанию product)
        &order_by=value|qty|name     # сортировка (по умолчанию value)
        &direction=desc|asc          # направление сортировки (по умолчанию desc)
        &limit=1000                  # ограничение строк (только для group_by=product)

    Возвращает:
      - по товарам: product_id, product_name, unit_type, category, qty, avg_price, value
      - по категориям (group_by=category): category_id, category_name, qty, value, avg_price
      - totals: общий qty и value
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # -------- параметры
        warehouse_id = request.query_params.get("warehouse")
        category_id = request.query_params.get("category")
        product_ids = request.query_params.getlist("product")
        include_zero = request.query_params.get("include_zero", "false").lower() == "true"
        group_by = request.query_params.get("group_by", "product")  # product|category
        order_by = request.query_params.get("order_by", "value")  # value|qty|name
        direction = request.query_params.get("direction", "desc")
        try:
            limit = int(request.query_params.get("limit", 1000))
        except ValueError:
            limit = 1000
        limit = max(1, min(limit, 10000))

        # -------- определяем "центральный" склад, если warehouse не задан
        warehouse = None
        if warehouse_id:
            warehouse = Warehouse.objects.filter(pk=warehouse_id).first()
        if not warehouse:
            # на крайний случай — первый склад в системе
            warehouse = Warehouse.objects.order_by("id").first()

        if not warehouse:
            return Response({"detail": "Склад не найден. Задайте ?warehouse=ID или создайте склад."}, status=400)

        # -------- базовый queryset
        qs = (
            WarehouseProduct.objects
            .select_related("product", "product__category")
            .filter(warehouse_id=warehouse.id)
        )
        if category_id:
            qs = qs.filter(product__category_id=category_id)
        if product_ids:
            qs = qs.filter(product_id__in=product_ids)
        if not include_zero:
            qs = qs.filter(count__gt=0)

        # выражения ORM
        line_value = ExpressionWrapper(
            F("count") * F("price"),
            output_field=DecimalField(max_digits=20, decimal_places=2),
        )
        zero_dec = Value(0, output_field=DecimalField(max_digits=20, decimal_places=2))

        # -------- агрегация
        if group_by == "category":
            grouped = (
                qs.values("product__category_id", "product__category__name")
                .annotate(
                    qty=Coalesce(Sum("count"), 0),
                    value=Coalesce(Sum(line_value), zero_dec),
                )
            )
            # средняя цена по категории = value / qty
            rows = []
            for r in grouped:
                qty = r["qty"] or 0
                value = r["value"] or Decimal("0")
                rows.append({
                    "category_id": r["product__category_id"],
                    "category_name": r["product__category__name"],
                    "qty": qty,
                    "value": value,
                    "avg_price": (value / qty) if qty else None,
                })
            # сортировка
            key_map = {
                "qty": (lambda x: x["qty"]),
                "value": (lambda x: x["value"]),
                "name": (lambda x: (x["category_name"] or "").lower()),
            }
            rows.sort(key=key_map.get(order_by, key_map["value"]), reverse=(direction != "asc"))

        else:  # group_by == "product"
            grouped = (
                qs.values(
                    "product_id",
                    "product__name",
                    "product__unit_type",
                    "product__category_id",
                    "product__category__name",
                )
                .annotate(
                    qty=Coalesce(Sum("count"), 0),
                    value=Coalesce(Sum(line_value), zero_dec),
                    # средняя цена = сумм(цена*кол-во)/сумм(кол-во)
                    weighted_price_num=Coalesce(Sum(line_value), zero_dec),
                    weighted_price_den=Coalesce(Sum("count"), 0),
                )
            )
            rows = []
            for r in grouped[:limit]:
                qty = r["qty"] or 0
                value = r["value"] or Decimal("0")
                avg_price = (r["weighted_price_num"] / r["weighted_price_den"]) if r["weighted_price_den"] else None
                rows.append({
                    "product_id": r["product_id"],
                    "product_name": r["product__name"],
                    "unit_type": r["product__unit_type"],
                    "category_id": r["product__category_id"],
                    "category_name": r["product__category__name"],
                    "qty": qty,
                    "value": value,
                    "avg_price": avg_price,
                })
            key_map = {
                "qty": (lambda x: x["qty"]),
                "value": (lambda x: x["value"]),
                "name": (lambda x: (x["product_name"] or "").lower()),
            }
            rows.sort(key=key_map.get(order_by, key_map["value"]), reverse=(direction != "asc"))

        # -------- тоталы
        totals_qs = qs.aggregate(
            sum_qty=Coalesce(Sum("count"), 0),
            sum_value=Coalesce(Sum(line_value), zero_dec),
        )
        totals = {
            "sum_qty": totals_qs["sum_qty"] or 0,
            "sum_value": totals_qs["sum_value"] or Decimal("0"),
        }

        return Response({
            "warehouse": {"id": warehouse.id, "name": warehouse.name},
            "group_by": group_by,
            "filters": {
                "category": category_id,
                "product_ids": product_ids or None,
                "include_zero": include_zero,
                "order_by": order_by,
                "direction": direction,
            },
            "totals": totals,
            "results": rows,
        })


class StocksByWarehouseDealerView(APIView):
    """
    Остатки по складам/дилерам
    GET /api/reports/stocks-by-warehouse-dealer/
        ?group_by=warehouse|dealer|warehouse_product   # по складам / по дилерам (responsible) / детально по товарам в складе
        &include_zero=false|true                       # по умолчанию false — показывать только позиции с qty > 0
        &category=ID                                   # фильтр по категории товара
        &product=ID&product=...                        # можно несколько product=
        &order_by=value|qty|name                       # ключ сортировки (для warehouse/dealer/warehouse_product)
        &direction=desc|asc                            # направление сортировки
        &limit=1000                                    # лимит строк (для детального разреза)

    Что отдаём:
      - group_by=warehouse:
          {warehouse_id, warehouse_name, qty, value, avg_price, skus, dealers}  # dealers — кол-во ответственных (обычно 1)
      - group_by=dealer:
          {dealer_id, dealer_name, qty, value, avg_price, warehouses, skus}
      - group_by=warehouse_product:
          {warehouse_id, warehouse_name, product_id, product_name, unit_type, category_id, category_name, qty, value, avg_price}

      totals: sum_qty, sum_value
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        group_by = request.query_params.get("group_by", "warehouse")  # warehouse|dealer|warehouse_product
        include_zero = request.query_params.get("include_zero", "false").lower() == "true"
        category_id = request.query_params.get("category")
        product_ids = request.query_params.getlist("product")

        order_by = request.query_params.get("order_by", "value")
        direction = request.query_params.get("direction", "desc")
        try:
            limit = int(request.query_params.get("limit", 1000))
        except ValueError:
            limit = 1000
        limit = max(1, min(limit, 10000))

        qs = (
            WarehouseProduct.objects
            .select_related("warehouse", "warehouse__responsible", "product", "product__category")
        )

        if category_id:
            qs = qs.filter(product__category_id=category_id)
        if product_ids:
            qs = qs.filter(product_id__in=product_ids)
        if not include_zero:
            qs = qs.filter(count__gt=0)

        # выражения ORM
        line_value = ExpressionWrapper(
            F("count") * F("price"),
            output_field=DecimalField(max_digits=20, decimal_places=2),
        )
        zero_dec = Value(0, output_field=DecimalField(max_digits=20, decimal_places=2))

        # Общие итоги для долей/контекста
        totals_qs = qs.aggregate(
            sum_qty=Coalesce(Sum("count"), 0),
            sum_value=Coalesce(Sum(line_value), zero_dec),
        )
        sum_qty = totals_qs["sum_qty"] or 0
        sum_value = totals_qs["sum_value"] or Decimal("0")

        rows = []

        if group_by == "dealer":
            grouped = (
                qs.values("warehouse__responsible_id",
                          "warehouse__responsible__first_name",
                          "warehouse__responsible__last_name")
                .annotate(
                    qty=Coalesce(Sum("count"), 0),
                    value=Coalesce(Sum(line_value), zero_dec),
                    warehouses=Count("warehouse_id", distinct=True),
                    skus=Count("product_id", distinct=True),
                )
            )
            # Python-метрики, сортировка
            tmp = []
            for r in grouped:
                dealer_id = r["warehouse__responsible_id"]
                fname = r["warehouse__responsible__first_name"] or ""
                lname = r["warehouse__responsible__last_name"] or ""
                display_name = (f"{fname} {lname}".strip() or (f"User {dealer_id}" if dealer_id else "—"))

                qty = r["qty"] or 0
                value = r["value"] or Decimal("0")
                avg_price = (value / qty) if qty else None

                tmp.append({
                    "dealer_id": dealer_id,
                    "dealer_name": display_name,
                    "qty": qty,
                    "value": value,
                    "avg_price": avg_price,
                    "warehouses": r["warehouses"],
                    "skus": r["skus"],
                })

            key_map = {
                "qty": (lambda x: x["qty"]),
                "value": (lambda x: x["value"]),
                "name": (lambda x: (x["dealer_name"] or "").lower()),
            }
            tmp.sort(key=key_map.get(order_by, key_map["value"]), reverse=(direction != "asc"))
            rows = tmp[:limit]

        elif group_by == "warehouse":
            grouped = (
                qs.values("warehouse_id", "warehouse__name")
                .annotate(
                    qty=Coalesce(Sum("count"), 0),
                    value=Coalesce(Sum(line_value), zero_dec),
                    skus=Count("product_id", distinct=True),
                    dealers=Count("warehouse__responsible_id", distinct=True),
                )
            )
            tmp = []
            for r in grouped:
                qty = r["qty"] or 0
                value = r["value"] or Decimal("0")
                avg_price = (value / qty) if qty else None
                tmp.append({
                    "warehouse_id": r["warehouse_id"],
                    "warehouse_name": r["warehouse__name"],
                    "qty": qty,
                    "value": value,
                    "avg_price": avg_price,
                    "skus": r["skus"],
                    "dealers": r["dealers"],
                })
            key_map = {
                "qty": (lambda x: x["qty"]),
                "value": (lambda x: x["value"]),
                "name": (lambda x: (x["warehouse_name"] or "").lower()),
            }
            tmp.sort(key=key_map.get(order_by, key_map["value"]), reverse=(direction != "asc"))
            rows = tmp[:limit]

        else:  # group_by == "warehouse_product" (детализация)
            grouped = (
                qs.values("warehouse_id", "warehouse__name",
                          "product_id", "product__name", "product__unit_type",
                          "product__category_id", "product__category__name")
                .annotate(
                    qty=Coalesce(Sum("count"), 0),
                    value=Coalesce(Sum(line_value), zero_dec),
                )
            )
            tmp = []
            for r in grouped[:limit]:
                qty = r["qty"] or 0
                value = r["value"] or Decimal("0")
                avg_price = (value / qty) if qty else None
                tmp.append({
                    "warehouse_id": r["warehouse_id"],
                    "warehouse_name": r["warehouse__name"],
                    "product_id": r["product_id"],
                    "product_name": r["product__name"],
                    "unit_type": r["product__unit_type"],
                    "category_id": r["product__category_id"],
                    "category_name": r["product__category__name"],
                    "qty": qty,
                    "value": value,
                    "avg_price": avg_price,
                })
            key_map = {
                "qty": (lambda x: x["qty"]),
                "value": (lambda x: x["value"]),
                "name": (lambda x: (x["product_name"] or "").lower()),
            }
            tmp.sort(key=key_map.get(order_by, key_map["value"]), reverse=(direction != "asc"))
            rows = tmp

        return Response({
            "group_by": group_by,
            "filters": {
                "include_zero": include_zero,
                "category": category_id,
                "product_ids": product_ids or None,
                "order_by": order_by,
                "direction": direction,
                "limit": limit if group_by != "warehouse" else None,
            },
            "totals": {
                "sum_qty": sum_qty,
                "sum_value": sum_value,
            },
            "results": rows,
        })


class ForecastShortagesView(APIView):
    """
    Прогнозируемые дефициты
    GET /api/reports/forecast-shortages/
        ?warehouse=ID                   # фильтр по складу
        &category=ID                    # фильтр по категории товара
        &product=ID&product=...         # несколько конкретных товаров
        &window_days=60                 # окно для расчёта средней дневной потребности (по умолчанию 60)
        &threshold_days=14              # минимальный запас покрытия в днях (по умолчанию 14)
        &include_incoming=true|false    # учитывать будущие приходы (Income pending/active), по умолчанию true
        &status_out=finished            # какой статус Outcome считать потреблением (по умолчанию finished)
        &min_total_usage=0              # отсечь «шум»: мин. суммарное потребление в окне
        &include_zero_demand=false      # включать позиции без спроса (по умолчанию false)
        &order_by=doc|days|rate|stock   # сортировка: дефицитность (по умолчанию 'days')
        &direction=asc|desc             # направление сортировки (по умолчанию asc)
        &limit=1000

    Возвращает список риск-дефицитных позиций с метриками:
      - warehouse_id/name, product_id/name/category/unit
      - stock_qty, incoming_qty, daily_rate, days_of_cover, depletion_date
      - recommended_qty (сколько докупить до threshold_days покрытия)
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # --- параметры
        warehouse_id = request.query_params.get("warehouse")
        category_id = request.query_params.get("category")
        product_ids = request.query_params.getlist("product")
        window_days = int(request.query_params.get("window_days", 60))
        threshold_days = int(request.query_params.get("threshold_days", 14))
        include_incoming = request.query_params.get("include_incoming", "true").lower() == "true"
        status_out = request.query_params.get("status_out", Outcome.Status.finished)
        min_total_usage = int(request.query_params.get("min_total_usage", 0))
        include_zero_demand = request.query_params.get("include_zero_demand", "false").lower() == "true"
        order_by = request.query_params.get("order_by", "days")  # days|rate|stock|doc
        direction = request.query_params.get("direction", "asc")
        try:
            limit = int(request.query_params.get("limit", 1000))
        except ValueError:
            limit = 1000
        limit = max(1, min(limit, 10000))

        today = timezone.now().date()
        date_from = today - timedelta(days=window_days)

        # --- 1) Текущие остатки по складу/товару
        stock_qs = WarehouseProduct.objects.select_related("warehouse", "product", "product__category")
        if warehouse_id:
            stock_qs = stock_qs.filter(warehouse_id=warehouse_id)
        if category_id:
            stock_qs = stock_qs.filter(product__category_id=category_id)
        if product_ids:
            stock_qs = stock_qs.filter(product_id__in=product_ids)

        stock_agg = (
            stock_qs.values(
                "warehouse_id", "warehouse__name",
                "product_id", "product__name", "product__unit_type",
                "product__category_id", "product__category__name",
            )
            .annotate(stock_qty=Coalesce(Sum("count"), 0))
        )
        # dict ключ: (w,p)
        stock_map = {}
        for r in stock_agg:
            key = (r["warehouse_id"], r["product_id"])
            stock_map[key] = {
                "warehouse_id": r["warehouse_id"],
                "warehouse_name": r["warehouse__name"],
                "product_id": r["product_id"],
                "product_name": r["product__name"],
                "unit_type": r["product__unit_type"],
                "category_id": r["product__category_id"],
                "category_name": r["product__category__name"],
                "stock_qty": r["stock_qty"] or 0,
            }

        # --- 2) Потребление (OutcomeItem) за окно
        out_qs = OutcomeItem.objects.select_related("outcome", "product")
        if warehouse_id:
            out_qs = out_qs.filter(outcome__warehouse_id=warehouse_id)
        if category_id:
            out_qs = out_qs.filter(product__category_id=category_id)
        if product_ids:
            out_qs = out_qs.filter(product_id__in=product_ids)
        if status_out:
            out_qs = out_qs.filter(outcome__status=status_out)

        usage_agg = (
            out_qs.filter(
                outcome__created__date__gte=date_from,
                outcome__created__date__lte=today,
            )
            .values("outcome__warehouse_id", "product_id")
            .annotate(total_usage=Coalesce(Sum("count"), 0))
        )
        usage_map = {(r["outcome__warehouse_id"], r["product_id"]): r["total_usage"] for r in usage_agg}

        # --- 3) Будущие приходы (IncomeItem) pending/active
        incoming_map = {}
        if include_incoming:
            inc_qs = IncomeItem.objects.select_related("income", "product")
            if warehouse_id:
                inc_qs = inc_qs.filter(income__warehouse_id=warehouse_id)
            if category_id:
                inc_qs = inc_qs.filter(product__category_id=category_id)
            if product_ids:
                inc_qs = inc_qs.filter(product_id__in=product_ids)
            inc_qs = inc_qs.filter(income__status__in=[Income.Status.pending, Income.Status.active])

            inc_agg = (
                inc_qs.values("income__warehouse_id", "product_id")
                .annotate(incoming_qty=Coalesce(Sum("count"), 0))
            )
            incoming_map = {(r["income__warehouse_id"], r["product_id"]): r["incoming_qty"] for r in inc_agg}

        # --- 4) Собираем позиции (только те, что есть на складе или были продажи/приходы в окне)
        keys = set(stock_map.keys()) | set(usage_map.keys()) | set(incoming_map.keys())

        rows = []
        for key in keys:
            w_id, p_id = key
            base = stock_map.get(key, {
                "warehouse_id": w_id, "warehouse_name": None,
                "product_id": p_id, "product_name": None,
                "unit_type": None, "category_id": None, "category_name": None,
                "stock_qty": 0
            })
            stock_qty = int(base["stock_qty"] or 0)
            used = int(usage_map.get(key, 0) or 0)
            incoming = int(incoming_map.get(key, 0) or 0)

            # фильтр «шума»
            if not include_zero_demand and used == 0:
                continue
            if used < min_total_usage:
                continue

            # среднесуточное потребление
            daily_rate = (Decimal(used) / Decimal(window_days)) if window_days > 0 else None

            # дни покрытия и дата исчерпания
            if daily_rate and daily_rate > 0:
                effective_stock = Decimal(stock_qty + incoming)
                days_of_cover = float(effective_stock / daily_rate)
                depletion_date = (today + timedelta(days=ceil(days_of_cover))).isoformat()
            else:
                days_of_cover = None
                depletion_date = None

            # рекомендация по закупке до порога покрытия
            if daily_rate and daily_rate > 0:
                target_qty = Decimal(threshold_days) * daily_rate
                deficit = target_qty - Decimal(stock_qty + incoming)
                recommended_qty = int(ceil(float(deficit))) if deficit > 0 else 0
            else:
                recommended_qty = 0

            # маркер дефицита
            is_short = False
            if daily_rate and daily_rate > 0:
                is_short = days_of_cover < threshold_days
            else:
                # нет спроса — дефицит не отмечаем (если include_zero_demand=false мы сюда не попадём)
                is_short = False

            if not is_short:
                # показывать только риск-дефициты; если нужно видеть все — добавим флаг позже
                continue

            rows.append({
                **base,
                "incoming_qty": incoming if include_incoming else 0,
                "window_days": window_days,
                "used_in_window": used,
                "daily_rate": float(daily_rate) if daily_rate is not None else None,
                "days_of_cover": round(days_of_cover, 2) if days_of_cover is not None else None,
                "depletion_date": depletion_date,
                "threshold_days": threshold_days,
                "recommended_qty": recommended_qty,
            })

        # --- сортировка
        def key_days(x):  # меньше дней покрытия — выше риск
            v = x["days_of_cover"]
            return (v is None, v if v is not None else float("inf"))

        def key_rate(x):  # больше расход — выше риск
            v = x["daily_rate"]
            return (v is None, -(v if v is not None else 0.0))

        def key_stock(x):
            return -x["stock_qty"]

        def key_doc(x):  # документальная срочность — дата исчерпания ближе
            v = x["depletion_date"]
            return (v is None, v or "9999-12-31")

        key_map = {
            "days": key_days,
            "rate": key_rate,
            "stock": key_stock,
            "doc": key_doc,
        }
        rows.sort(key=key_map.get(order_by, key_days), reverse=(direction == "desc"))

        return Response({
            "filters": {
                "warehouse": warehouse_id,
                "category": category_id,
                "product_ids": product_ids or None,
                "window_days": window_days,
                "threshold_days": threshold_days,
                "include_incoming": include_incoming,
                "status_out": status_out,
                "min_total_usage": min_total_usage,
                "include_zero_demand": include_zero_demand,
                "order_by": order_by, "direction": direction,
            },
            "count": min(len(rows), limit),
            "results": rows[:limit],
        })


class PlanVsActualView(APIView):
    """
    План продаж vs. Фактические продажи (по месяцам/дилерам/регионам)
    GET /api/reports/plan-vs-actual/
        ?date_from=YYYY-MM-DD
        &date_to=YYYY-MM-DD
        &dimension=none|dealer|warehouse     # разрез сравнения (по умолчанию dealer)
        &status=finished                     # какой статус Outcome считать фактом (по умолчанию finished)
        &client=ID&client=...                # фильтр по дилерам (для dealer/none)
        &warehouse=ID                        # фильтр по складу (для warehouse/none)
        &metric=amount|qty|orders            # ключ сортировки (по умолчанию amount)
        &direction=desc|asc
        &limit=200

    Логика:
      - Факт: OutcomeItem по периодам (TruncMonth(outcome.created)), сумма = count*price.
      - План: Report (status=confirmed), месяц берём из Report.period (1..12),
              год — из Report.created.year. Сумма = sum(ReportItem.count * Product.price).
      - При dimension=dealer план привязывается к Report.client.
      - При dimension=warehouse у плана нет разреза по складам -> плановые поля = null.

    Возвращаем по каждому месяцу и значению измерения (dimension_key):
      {
        "period": "YYYY-MM-01",
        "dimension": {"type": "...", "id": ..., "name": "..."},
        "actual_qty", "actual_amount", "actual_orders",
        "plan_qty",   "plan_amount",   "plan_orders",
        "diff_qty", "diff_amount", "diff_orders",
        "achv_amount_pct", "achv_qty_pct", "achv_orders_pct"
      }
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # -------- параметры
        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")
        dimension = request.query_params.get("dimension", "dealer")  # none|dealer|warehouse
        status = request.query_params.get("status", Outcome.Status.finished)
        client_ids = request.query_params.getlist("client")
        warehouse_id = request.query_params.get("warehouse")
        metric = request.query_params.get("metric", "amount")
        direction = request.query_params.get("direction", "desc")
        try:
            limit = int(request.query_params.get("limit", 200))
        except ValueError:
            limit = 200
        limit = max(1, min(limit, 2000))

        # -------- факт из OutcomeItem
        amount_expr = ExpressionWrapper(
            F("count") * F("price"),
            output_field=DecimalField(max_digits=20, decimal_places=2),
        )
        zero_dec = Value(0, output_field=DecimalField(max_digits=20, decimal_places=2))

        fact_qs = OutcomeItem.objects.select_related("outcome", "product", "outcome__client", "outcome__warehouse")
        if status:
            fact_qs = fact_qs.filter(outcome__status=status)
        if date_from:
            fact_qs = fact_qs.filter(outcome__created__date__gte=date_from)
        if date_to:
            fact_qs = fact_qs.filter(outcome__created__date__lte=date_to)
        if client_ids:
            fact_qs = fact_qs.filter(outcome__client_id__in=client_ids)
        if warehouse_id:
            fact_qs = fact_qs.filter(outcome__warehouse_id=warehouse_id)

        # поля группировки по измерению
        dim_values = []
        if dimension == "warehouse":
            dim_values = ["outcome__warehouse_id", "outcome__warehouse__name"]
        elif dimension == "dealer":
            dim_values = ["outcome__client_id", "outcome__client__first_name", "outcome__client__last_name"]
        else:
            dim_values = []  # none

        fact_grouped = (
            fact_qs.annotate(period=TruncMonth("outcome__created"))
            .values("period", *dim_values)
            .annotate(
                actual_qty=Coalesce(Sum("count"), 0),
                actual_amount=Coalesce(Sum(amount_expr), zero_dec),
                actual_orders=Count("outcome_id", distinct=True),
            )
            .order_by("period")
        )

        # -------- план из Report / ReportItem
        # report.period = месяц (1..12), report.created.year = год плана
        # план считаем в тех же границах дат (по дате создания отчёта)
        plan_rows = []
        if dimension in {"dealer", "none"}:
            plan_qs = Report.objects.filter(status=Report.Status.confirmed)
            if date_from:
                plan_qs = plan_qs.filter(created__date__gte=date_from)
            if date_to:
                plan_qs = plan_qs.filter(created__date__lte=date_to)
            if client_ids:
                plan_qs = plan_qs.filter(client_id__in=client_ids)

            # Присоединяем ReportItem и Product для цены
            plan_items = (
                ReportItem.objects.select_related("report", "product")
                .filter(report__in=plan_qs)
                .values(
                    "report__client_id",
                    "report__client__first_name",
                    "report__client__last_name",
                    "report__period",
                    "report__created__year",
                )
                .annotate(
                    plan_qty=Coalesce(Sum("count"), 0),
                    plan_amount=Coalesce(Sum(F("count") * F("product__price"),
                                             output_field=DecimalField(max_digits=20, decimal_places=2)), zero_dec),
                    plan_orders=Count("report_id", distinct=True),  # кол-во план-доков
                )
                .order_by("report__created__year", "report__period")
            )

            # нормализуем «period» к первому числу месяца конкретного года
            for r in plan_items:
                y = r["report__created__year"]
                m = r["report__period"]
                period_key = date(y, m, 1).isoformat()
                plan_rows.append({
                    "period": period_key,
                    "client_id": r["report__client_id"],
                    "first_name": r["report__client__first_name"],
                    "last_name": r["report__client__last_name"],
                    "plan_qty": r["plan_qty"],
                    "plan_amount": r["plan_amount"],
                    "plan_orders": r["plan_orders"],
                })

        # при dimension=warehouse плановых данных по складу нет — оставим план = None

        # -------- сшивка факта и плана
        # ключ: (period_iso, dimension_key)
        def dim_key_from_fact(row):
            if dimension == "warehouse":
                return ("warehouse", row.get("outcome__warehouse_id") or None)
            if dimension == "dealer":
                return ("dealer", row.get("outcome__client_id") or None)
            return ("none", None)

        def dim_obj_from_fact(row):
            if dimension == "warehouse":
                return {"type": "warehouse",
                        "id": row.get("outcome__warehouse_id"),
                        "name": row.get("outcome__warehouse__name")}
            if dimension == "dealer":
                cid = row.get("outcome__client_id")
                fname = row.get("outcome__client__first_name") or ""
                lname = row.get("outcome__client__last_name") or ""
                disp = (f"{fname} {lname}").strip() or (f"User {cid}" if cid else "—")
                return {"type": "dealer", "id": cid, "name": disp}
            return {"type": "all", "id": None, "name": "All dealers"}

        def dim_key_from_plan(row):
            if dimension == "dealer":
                return ("dealer", row["client_id"])
            return ("none", None)

        # индекс плана
        plan_map = {}
        for r in plan_rows:
            key = (r["period"],) + dim_key_from_plan(r)
            plan_map[key] = r

        # собираем итоговые строки
        bucket = {}
        for r in fact_grouped:
            p = r["period"]
            pkey = p.date().isoformat() if hasattr(p, "date") else p
            dkey = dim_key_from_fact(r)
            key = (pkey, dkey)

            item = bucket.get(key, {
                "period": pkey,
                "dimension": dim_obj_from_fact(r),
                "actual_qty": 0, "actual_amount": Decimal("0"), "actual_orders": 0,
                "plan_qty": None, "plan_amount": None, "plan_orders": None,
            })
            item["actual_qty"] += r["actual_qty"] or 0
            item["actual_amount"] += r["actual_amount"] or Decimal("0")
            item["actual_orders"] += r["actual_orders"] or 0
            bucket[key] = item

        # подмешиваем план
        for key, item in list(bucket.items()):
            period_iso, dkey = key[0], key[1]
            if dimension == "dealer":
                plan_key = (period_iso, "dealer", item["dimension"]["id"])
            else:
                plan_key = (period_iso, "none", None)
            pr = plan_map.get(plan_key)
            if pr:
                item["plan_qty"] = pr["plan_qty"]
                item["plan_amount"] = pr["plan_amount"]
                item["plan_orders"] = pr["plan_orders"]

        # если нужно — добавим строки, где есть план, но нет факта
        for pkey, _, _ in set(plan_map.keys()):
            # для dealer случая добавим ниже по ключам
            pass
        if dimension == "dealer":
            for pkey, _, cid in plan_map.keys():
                bkey = (pkey, ("dealer", cid))
                if bkey not in bucket:
                    pr = plan_map[(pkey, "dealer", cid)]
                    fname = pr["first_name"] or ""
                    lname = pr["last_name"] or ""
                    disp = (f"{fname} {lname}").strip() or (f"User {cid}" if cid else "—")
                    bucket[bkey] = {
                        "period": pkey,
                        "dimension": {"type": "dealer", "id": cid, "name": disp},
                        "actual_qty": 0, "actual_amount": Decimal("0"), "actual_orders": 0,
                        "plan_qty": pr["plan_qty"], "plan_amount": pr["plan_amount"], "plan_orders": pr["plan_orders"],
                    }
        elif dimension == "none":
            # агрегируем весь план по месяцу
            monthly_plan = {}
            for (pkey, _, _), pr in plan_map.items():
                mp = monthly_plan.setdefault(pkey, {"qty": 0, "amount": Decimal("0"), "orders": 0})
                mp["qty"] += pr["plan_qty"] or 0
                mp["amount"] += pr["plan_amount"] or Decimal("0")
                mp["orders"] += pr["plan_orders"] or 0
            for pkey, pr in monthly_plan.items():
                bkey = (pkey, ("none", None))
                if bkey not in bucket:
                    bucket[bkey] = {
                        "period": pkey,
                        "dimension": {"type": "all", "id": None, "name": "All dealers"},
                        "actual_qty": 0, "actual_amount": Decimal("0"), "actual_orders": 0,
                        "plan_qty": pr["qty"], "plan_amount": pr["amount"], "plan_orders": pr["orders"],
                    }

        # расчёт отклонений
        rows = []
        for v in bucket.values():
            aq, aa, ao = v["actual_qty"], v["actual_amount"], v["actual_orders"]
            pq, pa, po = v["plan_qty"], v["plan_amount"], v["plan_orders"]
            diff_qty = (aq - pq) if pq is not None else None
            diff_amount = (aa - pa) if pa is not None else None
            diff_orders = (ao - po) if po is not None else None
            achv_amount_pct = (float(aa) / float(pa) * 100.0) if pa not in (None, 0, Decimal("0")) else None
            achv_qty_pct = (float(aq) / float(pq) * 100.0) if pq not in (None, 0) else None
            achv_orders_pct = (float(ao) / float(po) * 100.0) if po not in (None, 0) else None

            rows.append({
                **v,
                "diff_qty": diff_qty,
                "diff_amount": diff_amount,
                "diff_orders": diff_orders,
                "achv_amount_pct": achv_amount_pct,
                "achv_qty_pct": achv_qty_pct,
                "achv_orders_pct": achv_orders_pct,
            })

        # сортировка + лимит
        key_map = {
            "amount": (lambda x: (x["actual_amount"] or Decimal("0"))),
            "qty": (lambda x: (x["actual_qty"] or 0)),
            "orders": (lambda x: (x["actual_orders"] or 0)),
        }
        rows.sort(key=key_map.get(metric, key_map["amount"]), reverse=(direction != "asc"))
        rows = rows[:limit]

        return Response({
            "dimension": dimension,
            "metric": metric,
            "filters": {
                "date_from": date_from,
                "date_to": date_to,
                "status": status,
                "client_ids": client_ids or None,
                "warehouse": warehouse_id,
            },
            "count": len(rows),
            "results": rows,
            "notes": [
                "План берётся из Report(status=confirmed), месяц = Report.period (1..12), год = Report.created.year.",
                "Сумма плана считается как ∑(ReportItem.count × Product.price).",
                "При dimension=warehouse плановые поля недоступны (нет привязки плана к складам)."
            ],
        })


class PlanAchievementView(APIView):
    """
    Выполнение плана в процентах
    GET /api/reports/plan-achievement/
        ?date_from=YYYY-MM-DD
        &date_to=YYYY-MM-DD
        &dimension=none|dealer                # разрез: общий или по дилерам (клиентам)
        &status=finished                      # статус Outcome для факта (по умолчанию finished)
        &client=ID&client=...                 # фильтр по дилерам (для dimension=dealer/none)
        &metric=amount|qty|orders             # чем сортировать строки (по умолчанию amount)
        &direction=desc|asc
        &limit=500
        &include_values=false|true            # добавить в ответ абсолютные план/факт (по умолчанию true)

    Возвращает по каждой строке (месяц × дилер или просто месяц):
      {
        "period": "YYYY-MM-01",
        "dimension": {"type": "dealer|all", "id": <id|None>, "name": "<display>"},
        "achv_amount_pct": 95.3,
        "achv_qty_pct": 88.4,
        "achv_orders_pct": 102.0,
        ...(опционально) "plan_amount", "actual_amount", ...
      }

    Примечания:
      - План: сумма = ∑(ReportItem.count × Product.price); qty = ∑ ReportItem.count; orders = кол-во Report.
      - Для разреза по складам планов нет — используйте /plan-vs-actual (dimension=warehouse) для факта.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # ---- параметры
        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")
        dimension = request.query_params.get("dimension", "none")  # none|dealer
        status = request.query_params.get("status", Outcome.Status.finished)
        client_ids = request.query_params.getlist("client")
        metric = request.query_params.get("metric", "amount")  # amount|qty|orders
        direction = request.query_params.get("direction", "desc")
        include_values = request.query_params.get("include_values", "true").lower() == "true"
        try:
            limit = int(request.query_params.get("limit", 500))
        except ValueError:
            limit = 500
        limit = max(1, min(limit, 2000))

        # ---- ФАКТ
        amount_expr = ExpressionWrapper(
            F("count") * F("price"),
            output_field=DecimalField(max_digits=20, decimal_places=2),
        )
        zero_dec = Value(0, output_field=DecimalField(max_digits=20, decimal_places=2))

        fact_qs = OutcomeItem.objects.select_related("outcome", "outcome__client")
        if status:
            fact_qs = fact_qs.filter(outcome__status=status)
        if date_from:
            fact_qs = fact_qs.filter(outcome__created__date__gte=date_from)
        if date_to:
            fact_qs = fact_qs.filter(outcome__created__date__lte=date_to)
        if client_ids:
            fact_qs = fact_qs.filter(outcome__client_id__in=client_ids)

        dim_fields = []
        if dimension == "dealer":
            dim_fields = [
                "outcome__client_id",
                "outcome__client__first_name",
                "outcome__client__last_name",
            ]

        fact_rows = (
            fact_qs.annotate(period=TruncMonth("outcome__created"))
            .values("period", *dim_fields)
            .annotate(
                actual_qty=Coalesce(Sum("count"), 0),
                actual_amount=Coalesce(Sum(amount_expr), zero_dec),
                actual_orders=Count("outcome_id", distinct=True),
            )
        )

        # ---- ПЛАН (Report.confirmed)
        plan_qs = Report.objects.filter(status=Report.Status.confirmed)
        if date_from:
            plan_qs = plan_qs.filter(created__date__gte=date_from)
        if date_to:
            plan_qs = plan_qs.filter(created__date__lte=date_to)
        if client_ids:
            plan_qs = plan_qs.filter(client_id__in=client_ids)

        plan_items = (
            ReportItem.objects.select_related("report", "product")
            .filter(report__in=plan_qs)
            .values(
                "report__client_id",
                "report__client__first_name",
                "report__client__last_name",
                "report__period",
                "report__created__year",
            )
            .annotate(
                plan_qty=Coalesce(Sum("count"), 0),
                plan_amount=Coalesce(
                    Sum(F("count") * F("product__price"),
                        output_field=DecimalField(max_digits=20, decimal_places=2)),
                    zero_dec
                ),
                plan_orders=Count("report_id", distinct=True),
            )
        )

        # ---- Индексы
        def month_iso(year: int, period_month: int) -> str:
            return date(year, period_month, 1).isoformat()

        plan_map = {}
        for r in plan_items:
            pkey = month_iso(r["report__created__year"], r["report__period"])
            if dimension == "dealer":
                dkey = ("dealer", r["report__client_id"])
                name = _display_name(
                    r["report__client__first_name"],
                    r["report__client__last_name"],
                    r["report__client_id"],
                )
            else:
                dkey = ("none", None)
                name = "All dealers"
            plan_map[(pkey, dkey)] = {
                "period": pkey,
                "dimension": {"type": dkey[0], "id": dkey[1], "name": name},
                "plan_qty": r["plan_qty"],
                "plan_amount": r["plan_amount"],
                "plan_orders": r["plan_orders"],
            }

        fact_map = {}
        for r in fact_rows:
            p = r["period"]
            pkey = p.date().isoformat() if hasattr(p, "date") else p
            if dimension == "dealer":
                cid = r.get("outcome__client_id")
                dkey = ("dealer", cid)
                name = _display_name(
                    r.get("outcome__client__first_name"),
                    r.get("outcome__client__last_name"),
                    cid,
                )
            else:
                dkey = ("none", None)
                name = "All dealers"

            item = fact_map.get((pkey, dkey), {
                "period": pkey,
                "dimension": {"type": dkey[0], "id": dkey[1], "name": name},
                "actual_qty": 0, "actual_amount": Decimal("0"), "actual_orders": 0,
            })
            item["actual_qty"] += r["actual_qty"] or 0
            item["actual_amount"] += r["actual_amount"] or Decimal("0")
            item["actual_orders"] += r["actual_orders"] or 0
            fact_map[(pkey, dkey)] = item

        # ---- Сшивка
        keys = set(plan_map.keys()) | set(fact_map.keys())
        rows = []
        for key in keys:
            pr = plan_map.get(key)
            fr = fact_map.get(key)

            period = (pr or fr)["period"]
            dim = (pr or fr)["dimension"]

            plan_qty = pr["plan_qty"] if pr else None
            plan_amount = pr["plan_amount"] if pr else None
            plan_orders = pr["plan_orders"] if pr else None

            actual_qty = fr["actual_qty"] if fr else 0
            actual_amount = fr["actual_amount"] if fr else Decimal("0")
            actual_orders = fr["actual_orders"] if fr else 0

            achv_amount_pct = (float(actual_amount) / float(plan_amount) * 100.0) \
                if plan_amount not in (None, 0, Decimal("0")) else None
            achv_qty_pct = (float(actual_qty) / float(plan_qty) * 100.0) \
                if plan_qty not in (None, 0) else None
            achv_orders_pct = (float(actual_orders) / float(plan_orders) * 100.0) \
                if plan_orders not in (None, 0) else None

            row = {
                "period": period,
                "dimension": dim,
                "achv_amount_pct": achv_amount_pct,
                "achv_qty_pct": achv_qty_pct,
                "achv_orders_pct": achv_orders_pct,
            }
            if include_values:
                row.update({
                    "plan_amount": plan_amount,
                    "actual_amount": actual_amount,
                    "plan_qty": plan_qty,
                    "actual_qty": actual_qty,
                    "plan_orders": plan_orders,
                    "actual_orders": actual_orders,
                })
            rows.append(row)

        # ---- сортировка
        def k_amount(x):
            return (x["achv_amount_pct"] is None, x["achv_amount_pct"] or -1.0)

        def k_qty(x):
            return (x["achv_qty_pct"] is None, x["achv_qty_pct"] or -1.0)

        def k_orders(x):
            return (x["achv_orders_pct"] is None, x["achv_orders_pct"] or -1.0)

        key_map = {"amount": k_amount, "qty": k_qty, "orders": k_orders}
        rows.sort(key=key_map.get(metric, k_amount), reverse=(direction != "asc"))

        rows = rows[:limit]

        return Response({
            "dimension": dimension,
            "metric": metric,
            "count": len(rows),
            "results": rows,
        })


def _display_name(first_name, last_name, cid):
    fname = first_name or ""
    lname = last_name or ""
    disp = (f"{fname} {lname}").strip() or (f"User {cid}" if cid else "—")
    return disp


class OrdersCountView(APIView):
    """
    Количество заказов в периоде
    GET /api/reports/orders-count/
        ?group_by=day|week|month          # по умолчанию: day
        &date_from=YYYY-MM-DD
        &date_to=YYYY-MM-DD
        &client=ID
        &warehouse=ID
        &status=pending|collected|...     # опционально: отфильтровать по конкретному статусу

    Возвращает список периодов с количеством заказов по статусам и итого.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        group_by = request.query_params.get("group_by", "day")
        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")
        client_id = request.query_params.get("client")
        warehouse_id = request.query_params.get("warehouse")
        status_filter = request.query_params.get("status")  # если задан — считаем только этот статус

        trunc_map = {"day": TruncDay, "week": TruncWeek, "month": TruncMonth}
        trunc = trunc_map.get(group_by, TruncDay)

        qs = Order.objects.all()
        if date_from:
            qs = qs.filter(created__date__gte=date_from)
        if date_to:
            qs = qs.filter(created__date__lte=date_to)
        if client_id:
            qs = qs.filter(client_id=client_id)
        if warehouse_id:
            qs = qs.filter(warehouse_id=warehouse_id)
        if status_filter:
            qs = qs.filter(status=status_filter)

        bucketed = (
            qs.annotate(period=trunc("created"))
            .values("period")
            .annotate(
                total=Coalesce(Count("id"), 0),
                pending=Coalesce(Count("id", filter=Q(status=Order.Status.pending)), 0),
                collected=Coalesce(Count("id", filter=Q(status=Order.Status.collected)), 0),
                delivering=Coalesce(Count("id", filter=Q(status=Order.Status.delivering)), 0),
                delivered=Coalesce(Count("id", filter=Q(status=Order.Status.delivered)), 0),
                sent=Coalesce(Count("id", filter=Q(status=Order.Status.sent)), 0),
                cancelled=Coalesce(Count("id", filter=Q(status=Order.Status.cancelled)), 0),
            )
            .order_by("period")
        )

        results = []
        summary = {
            "total": 0, "pending": 0, "collected": 0, "delivering": 0,
            "delivered": 0, "sent": 0, "cancelled": 0,
        }

        for row in bucketed:
            p = row["period"]
            key = p.date().isoformat() if hasattr(p, "date") else p

            item = {
                "period": key,
                "total": row["total"],
                "pending": row["pending"],
                "collected": row["collected"],
                "delivering": row["delivering"],
                "delivered": row["delivered"],
                "sent": row["sent"],
                "cancelled": row["cancelled"],
            }
            results.append(item)
            for k in summary.keys():
                summary[k] += item[k]

        return Response({
            "group_by": group_by,
            "filters": {
                "date_from": date_from, "date_to": date_to,
                "client": client_id, "warehouse": warehouse_id,
                "status": status_filter,
            },
            "summary": summary,
            "results": results,
        })


class AverageOrderAmountView(APIView):
    """
    Средняя сумма заказа
    GET /api/reports/average-order-amount/
        ?group_by=none|day|week|month        # временные корзины (по умолчанию month)
        ?dimension=none|dealer|warehouse     # дополнительный разрез (по умолчанию none)
        &date_from=YYYY-MM-DD
        &date_to=YYYY-MM-DD
        &status=finished|active|...          # статус Outcome (по умолчанию finished)
        &client=ID&client=...                # фильтр дилеров (для dimension=dealer/none)
        &warehouse=ID                        # фильтр склада (для dimension=warehouse/none)
        &order_by=avg|amount|orders          # сортировка строк (по умолчанию avg)
        &direction=desc|asc                  # направление сортировки (по умолчанию desc)
        &limit=500

    Возвращает для каждой строки:
      {
        "period": "YYYY-MM-01" | null (если group_by=none),
        "dimension": {"type": "dealer|warehouse|all", "id": <id|None>, "name": "<ФИО/склад|All>"},
        "orders": <int>,                    # количество документов Outcome
        "total_amount": "<decimal>",        # сумма продаж в периоде/разрезе
        "avg_order_amount": "<decimal>"     # средний чек = total_amount / orders
      }
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # ---- параметры
        group_by = request.query_params.get("group_by", "month")  # none|day|week|month
        dimension = request.query_params.get("dimension", "none")  # none|dealer|warehouse
        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")
        status = request.query_params.get("status", Outcome.Status.finished)
        client_ids = request.query_params.getlist("client")
        warehouse_id = request.query_params.get("warehouse")
        order_by = request.query_params.get("order_by", "avg")  # avg|amount|orders
        direction = request.query_params.get("direction", "desc")
        try:
            limit = int(request.query_params.get("limit", 500))
        except ValueError:
            limit = 500
        limit = max(1, min(limit, 2000))

        # ---- база
        qs = OutcomeItem.objects.select_related("outcome", "outcome__client", "outcome__warehouse")
        if status:
            qs = qs.filter(outcome__status=status)
        if date_from:
            qs = qs.filter(outcome__created__date__gte=date_from)
        if date_to:
            qs = qs.filter(outcome__created__date__lte=date_to)
        if client_ids:
            qs = qs.filter(outcome__client_id__in=client_ids)
        if warehouse_id:
            qs = qs.filter(outcome__warehouse_id=warehouse_id)

        # выражения ORM
        amount_expr = ExpressionWrapper(
            F("count") * F("price"),
            output_field=DecimalField(max_digits=20, decimal_places=2),
        )
        zero_dec = Value(0, output_field=DecimalField(max_digits=20, decimal_places=2))

        # поля группировки
        time_annot = None
        if group_by == "day":
            time_annot = TruncDay("outcome__created")
        elif group_by == "week":
            time_annot = TruncWeek("outcome__created")
        elif group_by == "month":
            time_annot = TruncMonth("outcome__created")

        values = []
        if time_annot is not None:
            qs = qs.annotate(period=time_annot)
            values.append("period")

        if dimension == "dealer":
            values += ["outcome__client_id", "outcome__client__first_name", "outcome__client__last_name"]
        elif dimension == "warehouse":
            values += ["outcome__warehouse_id", "outcome__warehouse__name"]

        # агрегация
        grouped = (
            qs.values(*values)
            .annotate(
                total_amount=Coalesce(Sum(amount_expr), zero_dec),
                orders=Count("outcome_id", distinct=True),
            )
        )

        # подготовка строк
        rows = []
        for r in grouped:
            # период
            if "period" in r:
                p = r["period"]
                period_key = p.date().isoformat() if hasattr(p, "date") else p
            else:
                period_key = None

            # измерение
            if dimension == "dealer":
                cid = r.get("outcome__client_id")
                fname = r.get("outcome__client__first_name") or ""
                lname = r.get("outcome__client__last_name") or ""
                name = (f"{fname} {lname}").strip() or (f"User {cid}" if cid else "—")
                dim_obj = {"type": "dealer", "id": cid, "name": name}
            elif dimension == "warehouse":
                wid = r.get("outcome__warehouse_id")
                wname = r.get("outcome__warehouse__name")
                dim_obj = {"type": "warehouse", "id": wid, "name": wname}
            else:
                dim_obj = {"type": "all", "id": None, "name": "All"}

            orders = r["orders"] or 0
            total_amount = r["total_amount"] or Decimal("0")
            avg_order_amount = (total_amount / orders) if orders else None

            rows.append({
                "period": period_key,
                "dimension": dim_obj,
                "orders": orders,
                "total_amount": total_amount,
                "avg_order_amount": avg_order_amount,
            })

        # сортировка
        def k_avg(x):
            return (x["avg_order_amount"] is None, x["avg_order_amount"] or Decimal("0"))

        def k_amount(x):
            return x["total_amount"] or Decimal("0")

        def k_orders(x):
            return x["orders"] or 0

        key_map = {"avg": k_avg, "amount": k_amount, "orders": k_orders}
        rows.sort(key=key_map.get(order_by, k_avg), reverse=(direction != "asc"))

        # лимит
        rows = rows[:limit]

        return Response({
            "group_by": group_by,
            "dimension": dimension,
            "filters": {
                "date_from": date_from, "date_to": date_to,
                "status": status, "client_ids": client_ids or None, "warehouse": warehouse_id,
            },
            "order_by": order_by, "direction": direction,
            "count": len(rows),
            "results": rows,
        })


class MostOrderedProductsView(APIView):
    """
    Самые часто заказываемые позиции
    GET /api/reports/most-ordered-products/
        ?date_from=YYYY-MM-DD
        &date_to=YYYY-MM-DD
        &client=ID
        &warehouse=ID
        &status=pending|collected|delivered|...   # фильтр по статусу заказа
        &metric=orders|qty|amount                 # сортировка (по умолчанию orders)
        &limit=50

    Возвращает список товаров с метриками:
      - product_id, product_name, unit_type, category
      - orders: в скольких заказах встречался товар
      - total_qty: общее количество по всем заказам
      - total_amount: сумма (qty × price)
      - share_orders_pct: доля товара в общем числе заказов
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        date_from = request.query_params.get("date_from")
        date_to = request.query_params.get("date_to")
        client_id = request.query_params.get("client")
        warehouse_id = request.query_params.get("warehouse")
        status = request.query_params.get("status")
        metric = request.query_params.get("metric", "orders")  # orders|qty|amount
        try:
            limit = int(request.query_params.get("limit", 50))
        except ValueError:
            limit = 50
        limit = max(1, min(limit, 500))

        qs = OrderItem.objects.select_related("order", "product", "product__category")
        if date_from:
            qs = qs.filter(order__created__date__gte=date_from)
        if date_to:
            qs = qs.filter(order__created__date__lte=date_to)
        if client_id:
            qs = qs.filter(order__client_id=client_id)
        if warehouse_id:
            qs = qs.filter(order__warehouse_id=warehouse_id)
        if status:
            qs = qs.filter(order__status=status)

        amount_expr = ExpressionWrapper(
            F("count") * F("price"),
            output_field=DecimalField(max_digits=20, decimal_places=2),
        )
        zero_dec = Value(0, output_field=DecimalField(max_digits=20, decimal_places=2))

        grouped = (
            qs.values("product_id", "product__name", "product__unit_type",
                      "product__category_id", "product__category__name")
            .annotate(
                total_qty=Coalesce(Sum("count"), 0),
                total_amount=Coalesce(Sum(amount_expr), zero_dec),
                orders=Count("order_id", distinct=True),
            )
        )

        # Общие заказы (для расчёта долей)
        total_orders = Order.objects.all()
        if date_from:
            total_orders = total_orders.filter(created__date__gte=date_from)
        if date_to:
            total_orders = total_orders.filter(created__date__lte=date_to)
        if client_id:
            total_orders = total_orders.filter(client_id=client_id)
        if warehouse_id:
            total_orders = total_orders.filter(warehouse_id=warehouse_id)
        if status:
            total_orders = total_orders.filter(status=status)
        total_orders_count = total_orders.count() or 0

        # сортировка
        order_key = {
            "qty": "-total_qty",
            "amount": "-total_amount",
            "orders": "-orders",
        }.get(metric, "-orders")
        grouped = grouped.order_by(order_key)

        results = []
        for i, r in enumerate(grouped[:limit], start=1):
            share_orders = (r["orders"] * 100.0 / total_orders_count) if total_orders_count else None
            results.append({
                "rank": i,
                "product_id": r["product_id"],
                "product_name": r["product__name"],
                "unit_type": r["product__unit_type"],
                "category_id": r["product__category_id"],
                "category_name": r["product__category__name"],
                "total_qty": r["total_qty"],
                "total_amount": r["total_amount"],
                "orders": r["orders"],
                "share_orders_pct": share_orders,
            })

        return Response({
            "metric": metric,
            "filters": {
                "date_from": date_from, "date_to": date_to,
                "client": client_id, "warehouse": warehouse_id,
                "status": status,
            },
            "total_orders": total_orders_count,
            "count": len(results),
            "results": results,
        })


class OrderImportView(APIView):
    """
    Импорт заказов из Excel-файла
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        excel_file = request.FILES.get("file")

        if not excel_file:
            return Response({"error": "Необходимо передать файл"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
        except Exception as e:
            return Response({"error": f"Ошибка при чтении файла: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

        created_count = 0
        skipped = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                phone_number = str(row[0]).strip()
                comment = row[1]
                total_amount = row[2]
                status_value = row[3]

                # ищем клиента по номеру телефона
                client = User.objects.filter(phone_number=phone_number).first()
                if not client:
                    skipped.append(f"Клиент с телефоном {phone_number} не найден")
                    continue

                Order.objects.create(
                    client=client,
                    user=request.user,  # авторизованный пользователь
                    comment=comment,
                    total_amount=total_amount or 0,
                    status=status_value or Order.Status.pending,
                )

                created_count += 1

            except Exception as e:
                skipped.append(str(e))
                continue

        return Response({
            "created": created_count,
            "skipped": len(skipped),
            "errors": skipped
        }, status=status.HTTP_201_CREATED)


class IncomeImportView(APIView):
    """
    Импорт приходов Income из Excel.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        excel_file = request.FILES.get("file")

        if not excel_file:
            return Response({"error": "Необходимо передать Excel-файл"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
        except Exception as e:
            return Response({"error": f"Ошибка при чтении файла: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

        created_count = 0
        skipped = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                phone = str(row[0]).strip()
                comment = row[1]
                total_amount = row[2]
                status_value = (row[3] or Income.Status.pending).strip() if row[3] else Income.Status.pending
                warehouse_name = str(row[4]).strip() if row[4] else None

                # 1️⃣ ищем клиента
                client = User.objects.filter(phone_number=phone).first()
                if not client:
                    skipped.append(f"Клиент с телефоном {phone} не найден")
                    continue

                # 2️⃣ ищем склад
                warehouse = None
                if warehouse_name:
                    warehouse = Warehouse.objects.filter(name__iexact=warehouse_name).first()
                    if not warehouse:
                        skipped.append(f"Склад '{warehouse_name}' не найден")
                        continue

                # 3️⃣ создаём Income
                Income.objects.create(
                    client=client,
                    user=request.user,
                    comment=comment or "",
                    total_amount=total_amount or 0,
                    status=status_value,
                    warehouse=warehouse,
                )
                created_count += 1

            except Exception as e:
                skipped.append(str(e))
                continue

        return Response({
            "created": created_count,
            "skipped": len(skipped),
            "errors": skipped
        }, status=status.HTTP_201_CREATED)


class BannerViewSet(viewsets.ModelViewSet):
    queryset = Banner.objects.order_by('-id')
    serializer_class = BannerSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPagination
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['status']


class CatalogViewSet(viewsets.ModelViewSet):
    queryset = Catalog.objects.order_by('-id')
    serializer_class = CatalogSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPagination
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['status']
