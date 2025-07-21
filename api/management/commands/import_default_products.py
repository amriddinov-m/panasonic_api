import os

import pandas as pd
from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from api.models import ProductCategory, Product  # замените на своё название приложения

class Command(BaseCommand):
    help = 'Импортирует категории и продукты из Excel-файла'

    def handle(self, *args, **kwargs):
        file_path = os.path.join(os.path.dirname(__file__), "Viko price.xlsx") # укажи путь, если файл не в корне проекта
        wb = load_workbook(filename=file_path)
        sheet = wb.active

        current_category = None

        for row in sheet.iter_rows(min_row=2):
            code_cell = row[1]  # колонка B
            name_cell = row[2]  # колонка C (в ней категории и названия товаров)
            price_cell = row[3]  # колонка D

            # Проверим, объединена ли ячейка name_cell
            for merged_range in sheet.merged_cells.ranges:
                if name_cell.coordinate in merged_range:
                    # Берем значение и стиль из верхней левой ячейки объединённого диапазона
                    top_left_cell = sheet[merged_range.coord.split(":")[0]]
                    name_cell = top_left_cell
                    break

            name = name_cell.value
            code = code_cell.value
            price = price_cell.value

            if not name:
                continue

            is_bold = bool(name_cell.font and name_cell.font.bold)

            self.stdout.write(f"Строка {name_cell.row}: '{name}' | bold={is_bold}")

            if is_bold:
                # Это категория
                category_name = str(name).strip()
                current_category, created = ProductCategory.objects.get_or_create(name=category_name)
                self.stdout.write(self.style.SUCCESS(
                    f"{'Создана' if created else 'Найдена'} категория: {current_category.name}"
                ))
                continue

            if not current_category:
                self.stdout.write(self.style.WARNING(f'Пропущен продукт без категории: {name}'))
                continue

            product_code = str(code).strip() if code else ''
            product_name = name.strip()

            try:
                product_price = round(float(price), 2)
            except (TypeError, ValueError):
                product_price = 0.0

            product, created = Product.objects.get_or_create(
                code=product_code,
                defaults={
                    'name': product_name,
                    'price': product_price,
                    'category': current_category,
                }
            )

            if not created:
                product.name = product_name
                product.price = product_price
                product.category = current_category
                product.save()

            self.stdout.write(self.style.NOTICE(
                f"{'Создан' if created else 'Обновлён'} продукт: {product.name} (код: {product.code})"
            ))

        self.stdout.write(self.style.SUCCESS("Импорт завершён."))
