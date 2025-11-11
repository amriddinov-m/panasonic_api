import os

import pandas as pd
from django.core.management.base import BaseCommand
from decimal import Decimal

from openpyxl import load_workbook

from api.models import Product, ProductCategory, WarehouseProduct


class Command(BaseCommand):
    help = 'Импорт остатков из Excel и создание WarehouseProduct'

    def handle(self, *args, **options):
        # путь к файлу рядом с командой
        file_path = os.path.join(os.path.dirname(__file__), "OSTATKA.xlsx")

        wb = load_workbook(filename=file_path)
        ws = wb.active  # активный лист

        # читаем заголовки из первой строки
        headers = [cell.value for cell in ws[1]]
        header_map = {name.strip().lower(): idx for idx, name in enumerate(headers) if name}

        # проверим, что нужные столбцы есть
        required = ["code", "name", "count"]
        for col in required:
            if col not in header_map:
                raise ValueError(f"❌ В Excel не найден столбец '{col}'")

        category, _ = ProductCategory.objects.get_or_create(name="test")

        created_products = 0
        updated_products = 0
        created_warehouse_records = 0

        # читаем данные со второй строки
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            code = row[header_map["code"]]
            name = row[header_map["name"]]
            count = row[header_map["count"]]


            if not code:
                self.stdout.write(self.style.WARNING(f"⚠️ Пропущена строка {i} без кода"))
                continue

            # безопасно приводим типы
            code = str(code).strip()
            name = str(name).strip() if name else code
            count = int(count) if count not in (None, "") else 0
            price = 0

            # ищем или создаём продукт
            product, created = Product.objects.get_or_create(
                code=code,
                defaults={
                    "name": name,
                    "category": category,
                    "user_id": 1,
                    "price": price,
                    "status": "active",
                },
            )

            if not created:
                updated_products += 1
            else:
                created_products += 1

            # создаём WarehouseProduct
            WarehouseProduct.objects.create(
                product=product,
                warehouse_id=2,
                count=count,
                status="active",
                price=price,
                user_id=1,
            )

            created_warehouse_records += 1

        self.stdout.write(self.style.SUCCESS(
            f"✅ Импорт завершён: {created_products} новых продуктов, "
            f"{updated_products} обновлено, {created_warehouse_records} складских записей создано."
        ))