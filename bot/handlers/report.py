import datetime
import io
import mimetypes
import re
import pandas as pd
import os
import tempfile
from mimetypes import MimeTypes

from aiogram import F, types
from aiogram.fsm.context import FSMContext
from aiogram.types import FSInputFile, Document, InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.utils.keyboard import InlineKeyboardBuilder
from django.core.files.base import ContentFile
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import Workbook
from openpyxl.styles import Side, Border, Alignment, Font
from openpyxl.utils import get_column_letter

from api.models import Report, WarehouseProduct
from bot.MESSAGES import MESSAGES
from bot.handlers.helpers import create_report_item
# from bot.handlers.helpers import add_days_to_today, file_processing, create_excel_task_file
# from bot.handlers.lot import TIME_ZONE
from bot.keyboards.main import main_menu_btn, inline_btns, task_btns
from bot.loader import dp, bot, form_router
# from bot.models import EEUser, Task
# from bot.state.task import TaskCreateState, TaskCompleteState

from bot.state.document import UploadState
from panasonic_api import settings
from user.models import User


# from ee_task.settings import ITEMS_PER_PAGE


@form_router.message(F.text == '📃 Шаблон отчёта')
async def create_task_step(message: types.Message, state: FSMContext):
    file_path = os.path.join(os.path.dirname(__file__), "reports/xlsx", "panasonic_default_sales_data.xlsx")
    file = FSInputFile(file_path)
    await message.answer_document(file, caption=MESSAGES['send_sales_report'])
    await state.set_state(UploadState.send_document)
    user = User.objects.get(tg_id=message.from_user.id)
    await state.update_data(user_id=user.id)


@form_router.message(F.document)
async def handle_excel_file(message: types.Message, state: FSMContext):
    if not message.document.file_name.endswith(".xlsx"):
        await message.answer("Пожалуйста, отправьте Excel-файл в формате .xlsx")
        return

    # Скачиваем файл во временное хранилище
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        file_path = tmp.name
        await message.bot.download(message.document, destination=file_path)

    try:
        df = pd.read_excel(file_path)

        if not {"Item Code", "Item Description", "Count"}.issubset(df.columns):
            await message.answer("❌ В файле отсутствуют необходимые колонки: Item Code, Item Description, Count")
            return

        items = []
        for _, row in df.iterrows():
            item_code = row["Item Code"]
            item_name = row["Item Description"]
            count = row["Count"]

            if pd.isna(item_code) or pd.isna(count):
                continue

            try:
                count = int(count)
            except ValueError:
                continue

            if count > 0:
                items.append({
                    "item_code": str(item_code).strip(),
                    "item_name": str(item_name).strip(),
                    "count": count
                })
        if not items:
            await message.answer("Файл прочитан, но не найдено ни одного товара с количеством > 0.")
            return

        await state.update_data(items=items)
        await state.set_state(UploadState.waiting_for_confirmation)

        preview_text = "\n".join(
            [f"📦 {item['item_name']} (Код: {item['item_code']}): {item['count']} шт." for item in items]
        )

        keyboard = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="✅ Всё верно", callback_data="confirm_data")],
            [InlineKeyboardButton(text="🔄 Отправить заново", callback_data="cancel_upload")]
        ])

        await message.answer(
            f"Проверьте данные:\n\n{preview_text}",
            reply_markup=keyboard
        )

    except Exception as e:
        await message.answer(f"❌ Ошибка при обработке файла: {e}")
    finally:
        os.remove(file_path)


@form_router.callback_query(F.data == "confirm_data")
async def confirm_upload(callback: types.CallbackQuery, state: FSMContext):
    data = await state.get_data()
    items = data.get("items", [])
    report = Report.objects.create(client_id=data.get('user_id'))
    for item in items:
        await create_report_item(report, item["item_code"], item["item_name"], item["count"])

    await state.clear()
    await callback.message.edit_reply_markup()
    await callback.message.answer("✅ Расходы по складу успешно зарегистрированы!")


@form_router.callback_query(F.data == "cancel_upload")
async def cancel_upload(callback: types.CallbackQuery, state: FSMContext):
    await state.clear()
    await callback.message.edit_reply_markup()
    await callback.message.answer("❌ Загрузка отменена. Пожалуйста, отправьте новый Excel-файл.")




@form_router.message(F.text == "📤 Выгрузить склад")
async def export_warehouse_products(message: types.Message):
    queryset = WarehouseProduct.objects.select_related("product").all()

    data = []
    for wp in queryset:
        data.append({
            "Код модели": wp.product.code,
            "Модели": wp.product.name,
            "Количество": wp.count
        })

    # Создаем Excel-файл в памяти
    wb = Workbook()
    ws = wb.active
    ws.title = "Warehouse"

    # Заголовки
    headers = list(data[0].keys()) if data else ["Код модели", "Модели", "Количество"]
    ws.append(headers)

    # Стили для заголовков
    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    # Данные
    for row_idx, row in enumerate(data, start=2):
        for col_idx, key in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row[key])
            cell.alignment = center_align
            cell.border = border

    # Автоширина колонок
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length + 2

    # Сохраняем файл в память
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Отправляем файл
    excel_file = types.BufferedInputFile(buffer.read(), filename="warehouse_export.xlsx")
    await message.answer_document(excel_file, caption="📦 Текущий список товаров на складе")



@form_router.message(F.text == '🗂 Прайс каталог')
async def create_task_step(message: types.Message):
    file_path = os.path.join(os.path.dirname(__file__), "reports/xlsx", "price_catalog.xlsx")
    file = FSInputFile(file_path)
    await message.answer_document(file, caption=MESSAGES['send_price_catalog'])


# @form_router.message(F.text == '📌 Создать поручение')
# async def create_task_step(message: types.Message, state: FSMContext):
#     if message.from_user.id in [settings.PROCUREMENT_TG_ID, settings.CONTRACT_TG_ID, settings.DIRECTOR_TG_ID]:
#         keyboard = main_menu_btn()
#         await bot.send_message(message.from_user.id,
#                                MESSAGES['task_title'],
#                                reply_markup=keyboard)
#         await state.set_state(TaskCreateState.title)
#     else:
#         await message.answer('❌ Действие запрещено!')
#
#
# @form_router.message(TaskCreateState.title)
# async def task_title_step(message: types.Message, state: FSMContext):
#     qs_dict = {}
#     if message.from_user.id == settings.DIRECTOR_TG_ID:
#         qs_dict['user_type'] = 'task'
#     elif message.from_user.id == settings.PROCUREMENT_TG_ID:
#         qs_dict['user_type'] = 'procurement'
#     else:
#         qs_dict['user_type'] = 'contract'
#     executors = EEUser.objects.exclude(tg_id=message.from_user.id).filter(**qs_dict)
#     markup = inline_btns(executors, 'executor_')
#     creator = EEUser.objects.get(tg_id=message.from_user.id)
#     if message.photo or message.document:
#         task = Task.objects.create(title=message.caption, creator=creator)
#         if message.photo:
#             file_id = message.photo[-1].file_id
#             file_unique_id = message.photo[-1].file_unique_id
#             file_info = await bot.get_file(file_id)
#             file_path = file_info.file_path
#             downloaded_file = await bot.download_file(file_path)
#             file_name = f"{file_unique_id}.jpg"
#             django_file = SimpleUploadedFile(file_name, downloaded_file.read(), content_type='image/jpeg')
#             task.file_id = file_id
#             task.file_type = 'photo'
#             task.file.save(file_name, django_file, save=True)
#         else:
#             file_id = message.document.file_id
#             file_info = await bot.get_file(file_id)
#             file_path = file_info.file_path
#             downloaded_file = await bot.download_file(file_path)
#             mime = MimeTypes()
#             mime_type, _ = mime.guess_type(message.document.file_name)
#
#             extension = mimetypes.guess_extension(mime_type)
#             extension = extension if extension else '.txt'
#             dynamic_file_name = f"{file_info.file_unique_id}{extension}"
#             task.file_id = file_id
#             task.file_type = 'document'
#             task.file.save(dynamic_file_name, ContentFile(downloaded_file.read()), save=True)
#         task.save()
#     else:
#         task = Task.objects.create(title=message.text, creator=creator)
#
#     sent_message = await bot.send_message(message.from_user.id,
#                                           MESSAGES['task_created'].format(task.pk), reply_markup=markup)
#     await state.update_data(task_id=task.pk, message_id=sent_message.message_id)
#     await state.set_state(TaskCreateState.executor)
#
#
# @form_router.callback_query(TaskCreateState.executor, F.data.startswith("executor_"))
# async def task_choose_executor_step(callback_query: types.CallbackQuery, state: FSMContext):
#     _, executor_id = callback_query.data.split('_')
#     await state.update_data(executor_id=executor_id)
#     await bot.edit_message_text(MESSAGES['period_of_execution'], callback_query.from_user.id,
#                                 callback_query.message.message_id, reply_markup=None)
#     await state.set_state(TaskCreateState.period)
#
#
# @form_router.message(TaskCreateState.executor)
# async def task_search_executor_step(message: types.Message, state: FSMContext):
#     executors = EEUser.objects.filter(fullname__contains=message.text)
#     markup = inline_btns(executors, 'executor_')
#     data = await state.get_data()
#     await bot.edit_message_text(text=MESSAGES['task_created'].format(data['task_id']), chat_id=message.from_user.id,
#                                 message_id=data['message_id'], reply_markup=markup)
#     await bot.delete_message(message.from_user.id, message.message_id)
#
#
# @form_router.message(TaskCreateState.period)
# async def period_of_execution_step(message: types.Message, state: FSMContext):
#     period = message.text
#     if re.match(r'^\d+$', period):
#         data = await state.get_data()
#         result_date = add_days_to_today(period)
#         task = Task.objects.get(id=data['task_id'])
#         task.executor_id = data['executor_id']
#         task.deadline = result_date
#         task.status = 'progress'
#         task.save()
#         executor = EEUser.objects.get(id=data['executor_id'])
#         builder = InlineKeyboardBuilder()
#         builder.row(
#             types.InlineKeyboardButton(
#                 text='✅Выполнено',
#                 callback_data=f'complete-task_{task.pk}'),
#         )
#
#         if task.file_id:
#             match task.file_type:
#                 case 'photo':
#                     await bot.send_photo(executor.tg_id, task.file_id, caption='📎 Прикреплённый файл')
#                 case 'document':
#                     await bot.send_document(executor.tg_id, task.file_id, caption='📎 Прикреплённый файл')
#         await bot.send_message(executor.tg_id, MESSAGES['executor_task']
#                                .format(id=task.pk, creator=task.creator.fullname, title=task.title,
#                                        status=task.get_status_display(), deadline=task.deadline),
#                                reply_markup=builder.as_markup())
#         keyboard = main_menu_btn()
#         await bot.send_message(message.from_user.id, MESSAGES['task_sent'], reply_markup=keyboard)
#         await state.clear()
#     else:
#         await message.answer("Пожалуйста, введите только цифры.")
#
#
# @form_router.callback_query(F.data.startswith('complete-task_'))
# async def complete_task_step(callback_query: types.CallbackQuery, state: FSMContext):
#     _, task_id = callback_query.data.split('_')
#     await callback_query.message.answer(MESSAGES['complete_task'])
#     await state.update_data(task_id=task_id, message_id=callback_query.message.message_id)
#     await state.set_state(TaskCompleteState.result_title)
#
#
# @form_router.message(TaskCompleteState.result_title)
# async def task_complete_title_step(message: types.Message, state: FSMContext):
#     data = await state.get_data()
#     file_data = file_processing(message)
#     task = Task.objects.get(id=data['task_id'])
#     task.result = file_data['title']
#     task.result_file_id = file_data['file_id']
#     task.result_file_type = file_data['file_type']
#     task.status = 'completed'
#     task.finish_date = datetime.datetime.now()
#     task.save()
#     keyboard = main_menu_btn()
#     await bot.send_message(message.from_user.id, MESSAGES['task_sent'], reply_markup=keyboard)
#     message_text = MESSAGES['executor_task'].format(id=task.pk, creator=task.creator.fullname,
#                                                     title=task.title,
#                                                     status=task.get_status_display(),
#                                                     deadline=task.deadline)
#     completed_message_text = MESSAGES['completed_task'].format(id=task.pk, executor=task.executor.fullname,
#                                                                result=task.result,
#                                                                created=task.created.astimezone(TIME_ZONE)
#                                                                .strftime("%d.%m.%Y | %H:%M"),
#                                                                status=task.get_status_display(),
#                                                                finish_date=task.finish_date
#                                                                .astimezone(TIME_ZONE).strftime("%d.%m.%Y | %H:%M"))
#     if task.result_file_id:
#         match task.result_file_type:
#             case 'photo':
#                 await bot.send_photo(task.creator.tg_id, task.result_file_id, caption='📎 Прикреплённый файл')
#             case 'document':
#                 await bot.send_document(task.creator.tg_id, task.result_file_id, caption='📎 Прикреплённый файл')
#     await bot.send_message(task.creator.tg_id, completed_message_text)
#     await bot.edit_message_text(message_text, message.from_user.id, int(data['message_id']), reply_markup=None)
#
#
# async def show_current_page(chat_id, data, update=False, message_id=0):
#     start_index = (current_page - 1) * ITEMS_PER_PAGE
#     end_index = start_index + ITEMS_PER_PAGE
#     current_data = data[start_index:end_index]
#     builder = InlineKeyboardBuilder()
#     total_pages = (len(data) + ITEMS_PER_PAGE - 1) // ITEMS_PER_PAGE
#     builder.row(
#         types.InlineKeyboardButton(text="⬅️", callback_data="pagination_prev"),
#         types.InlineKeyboardButton(text=f"{current_page}/{total_pages}", callback_data="current_page"),
#         types.InlineKeyboardButton(text="➡️", callback_data="pagination_next"),
#     )
#     text = f"Результаты {start_index}-{end_index} из {len(data)}:\n"
#     counter = start_index + 1
#     for task in current_data:
#         executor = '-'
#         if task.executor:
#             executor = task.executor.fullname
#         finish_date = '-'
#         if task.finish_date:
#             finish_date = task.finish_date.astimezone(TIME_ZONE).strftime("%d.%m.%Y | %H:%M")
#         text += (f'{counter}.'
#                  f'#️⃣ Поручение №: {task.pk}\n'
#                  f'👤 Исполнитель: {executor}\n'
#                  f'📅 Дата поручения: {task.created.astimezone(TIME_ZONE).strftime("%d.%m.%Y | %H:%M")}\n'
#                  f'📅 Дата выполнения: {finish_date}\n'
#                  f'🔄 Статус: {task.get_status_display()}\n\n')
#         counter += 1
#     if update:
#         await bot.edit_message_text(chat_id=chat_id, message_id=message_id, text=text, reply_markup=builder.as_markup())
#     else:
#         await bot.send_message(chat_id, text,
#                                reply_markup=builder.as_markup())
#
# current_page = 1
#
#
# @form_router.message(F.text == '📌 Мои поручения')
# async def my_tasks(message: types.Message, state: FSMContext):
#     user = EEUser.objects.get(tg_id=message.from_user.id)
#     qs_dict = {}
#     if message.from_user.id in [settings.PROCUREMENT_TG_ID, settings.CONTRACT_TG_ID, settings.DIRECTOR_TG_ID]:
#         qs_dict['creator'] = user
#     else:
#         qs_dict['executor'] = user
#     tasks = Task.objects.filter(**qs_dict)
#     global current_page
#     current_page = 1
#     create_excel_task_file(message.from_user.id)
#     await show_current_page(message.chat.id, tasks)
#     await state.update_data(_type='my_tasks')
#
#
# @dp.callback_query(F.data.startswith('pagination_'))
# async def handle_pagination_buttons(callback_query: types.CallbackQuery, state: FSMContext):
#     print(callback_query)
#     try:
#         global current_page
#         state_data = await state.get_data()
#         qs_dict = {}
#         user = EEUser.objects.get(tg_id=callback_query.from_user.id)
#         if callback_query.from_user.id in [settings.PROCUREMENT_TG_ID, settings.CONTRACT_TG_ID, settings.DIRECTOR_TG_ID]:
#             qs_dict['creator'] = user
#         else:
#             qs_dict['executor'] = user
#         data = Task.objects.filter(**qs_dict)
#         if state_data.get('_type') == 'overdue_tasks':
#             data = Task.get_overdue_tasks().filter(creator=user)
#
#         if callback_query.data == "pagination_prev" and current_page > 1:
#             current_page -= 1
#         elif callback_query.data == "pagination_next" and current_page < len(data) / ITEMS_PER_PAGE:
#             current_page += 1
#         await show_current_page(callback_query.message.chat.id, data, update=True,
#                                 message_id=callback_query.message.message_id)
#     except Exception as err:
#         print(err)
#
#
# @form_router.message(F.text == '🗓 Просроченные поручения')
# async def overdue_tasks_step(message: types.Message, state: FSMContext):
#     creator = EEUser.objects.get(tg_id=message.from_user.id)
#     tasks = Task.get_overdue_tasks().filter(creator=creator)
#     global current_page
#     current_page = 1
#     await show_current_page(message.chat.id, tasks)
#     await state.update_data(_type='overdue_tasks')
