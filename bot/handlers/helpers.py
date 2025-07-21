import datetime
import locale
import io

import requests
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from api.models import Product, ReportItem

# from bot.admin import url_file_download
# from bot.handlers.lot import TIME_ZONE
# from bot.models import Task, EEUser
# from ee_task import settings

locale.setlocale(locale.LC_NUMERIC, 'ru_RU.utf8')


def clear_int_string(int_string):
    """ Приводит значения вида ` 300 000` в `300000` """
    return int_string.strip().replace(' ', '')


def to_locale(value):
    return locale.format('%d', value, grouping=True)


def add_days_to_today(days):
    today = datetime.datetime.now().date()
    result_date = today + datetime.timedelta(days=int(days))
    return result_date


def file_processing(message):
    context = {}
    title = message.text
    file_id = ''
    file_type = ''
    if message.photo:
        file_id = message.photo[-1].file_id
        title = message.caption
        file_type = 'photo'
    elif message.document:
        file_id = message.document.file_id
        title = message.caption
        file_type = 'document'
    context['title'] = title
    context['file_id'] = file_id
    context['file_type'] = file_type
    return context


def set_column_width_based_on_content(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass

        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width


async def create_report_item(report, item_code: str, item_name: str, count: int):
    product = Product.objects.get(code=item_code)
    ReportItem.objects.create(report=report, product=product, count=count)

# def create_excel_task_file(chat_id):
#     file_path = f'C:\\Projects\\ee_task\\bot\\handlers\\reports\\xlsx\\data.xlsx'
#     creator = EEUser.objects.get(tg_id=chat_id)
#     tasks = Task.objects.filter(creator=creator)
#
#     wb = Workbook()
#     ws = wb.active
#
#     ws.append(['№', 'Задача', 'Исполнитель', 'Файл постановления', 'Файл выполнения', 'Дата постановления',
#                'Срок выполнения', 'Дата выполнения', 'Статус'])
#
#     for task in tasks:
#         file_id = '-'
#         result_file_id = '-'
#         if task.file_id or task.result_file_id:
#             file_id = '-'
#             if task.file_id:
#                 file_id = url_file_download(task.file_id)
#             elif task.result_file_id:
#                 result_file_id = url_file_download(task.result_file_id)
#
#             cell = ws.cell(row=ws.max_row, column=4)
#             cell.value = file_id
#             cell.hyperlink = f"#B{ws.max_row}"
#             cell.font = Font(underline='single')
#         executor = '-'
#         if task.executor:
#             executor = task.executor.fullname
#         finish_date = '-'
#         if task.finish_date:
#             finish_date = task.finish_date.astimezone(TIME_ZONE).strftime("%d.%m.%Y | %H:%M")
#         created = task.created.astimezone(TIME_ZONE).strftime("%d.%m.%Y | %H:%M")
#         ws.append([task.id, task.title, executor, file_id, result_file_id, f'{created}',
#                    task.deadline, f'{finish_date}', task.get_status_display()])
#
#     set_column_width_based_on_content(ws)
#
#     wb.save(file_path)
#
#     files = {'document': open(file_path, 'rb')}
#     data = {
#         'chat_id': chat_id,
#         'caption': f'📃 Выгрузка\n'
#     }
#     url = f'https://api.telegram.org/bot{settings.BOT_TOKEN}/sendDocument'
#
#     requests.post(url, data=data, files=files)

